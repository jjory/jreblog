# rebuild trigger v3
"""
app.py
─────────────────────────────────────────────────────────
JRE일본부동산 — 마이소크 → 네이버 블로그 자동 작성 시스템

특징:
- 한 번에 최대 5개 도면(마이소크) 업로드 → 블로그 5개 일괄 생성 (병렬 처리)
- JPG/PNG/WEBP/GIF/PDF 지원
- 파일마다 글 스타일을 따로 선택 가능
- 사무실 공용 비밀번호 인증 (작성자 이름 입력 없음, 새로고침해도 유지)
- 카카오톡 요약 원클릭 복사
- 생성된 블로그 전체 ZIP 다운로드 (G드라이브에 수동 저장)

실행:
- 로컬: streamlit run app.py  (.env)
- 클라우드: Streamlit Secrets
"""

import hashlib
import io
import json
import os
import re
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import NamedTemporaryFile
from zoneinfo import ZoneInfo

# 도쿄 타임존 — Drive 자동 처리 영역의 모든 시간 표시·로그에 사용
# (Render 서버는 UTC라 datetime.now()를 그대로 쓰면 9시간 차이 발생)
TZ_TOKYO = ZoneInfo("Asia/Tokyo")

import streamlit as st
from dotenv import load_dotenv

from src.analyzer import analyze_property_sheet, format_error_korean
from src.generator import (
    VISA_LABELS,
    build_naver_smarteditor_html,
    generate_blog_post,
    list_available_styles,
)
from src.persistence import (
    load_history,
    add_to_history,
    delete_from_history,
    clear_history,
    toggle_favorite,
    delete_old_history,
    save_session,
    load_session,
    clear_session,
    cleanup_old_sessions,
    generate_session_id,
    HISTORY_RETENTION_DAYS,
)

# 物件DB(제안용 리스트) — Postgres 저장. 미설정/미설치여도 앱은 정상 동작.
try:
    from src import db as property_db
except Exception:
    property_db = None

# 路線 시트 소요시간 조회 — 미설정/실패여도 앱은 정상 동작.
try:
    from src import sheet_sync
except Exception:
    sheet_sync = None

# 항목 한글 매핑(비자·構造·方向·設備) — 없으면 원문 표시로 폴백.
try:
    from src import field_map
except Exception:
    field_map = None

# 노선·역·지명 한글 번역 — 없으면 원문 표시로 폴백.
try:
    from src import translation_db
except Exception:
    translation_db = None

# ─────────────────────────────────────────────────
# 설정 로드 (로컬 .env / 클라우드 Streamlit Secrets)
# ─────────────────────────────────────────────────
load_dotenv()
try:
    for _key in st.secrets:
        _val = st.secrets[_key]
        if isinstance(_val, str):
            os.environ.setdefault(_key, _val)
except Exception:
    pass

MAX_UPLOADS = 5

# ─────────────────────────────────────────────────
# 동시 병렬 처리 워커 수
# ─────────────────────────────────────────────────
# 2로 설정 — Streamlit Cloud 무료 플랜(RAM 1GB) 안정성 최우선.
# - 단일 사용자: 5파일을 2+2+1 배치로 처리 (약 60초)
# - 3대 동시 사용: 3 × 2 = 6개 동시 처리 (메모리 여유)
# 속도보다 안정성을 우선. 큰 PDF가 섞여도 안정 작동.
MAX_PARALLEL_WORKERS = 5

st.set_page_config(
    page_title="🏠 JRE일본부동산 블로그 자동작성",
    page_icon="🏠",
    layout="wide",
)

# ─────────────────────────────────────────────────
# 📁 생성된 블로그 ZIP을 저장할 권장 폴더 경로
#     변경하려면 아래 한 줄만 바꾸세요.
# ─────────────────────────────────────────────────
OUTPUT_FOLDER_PATH = (
    r"G:\내 드라이브\0.사내공유\1.부동산_공유\1.안건\5.매물취합\블로그작성"
)


# ─────────────────────────────────────────────────
# 통계 분석 헬퍼 — 이력 데이터에서 정보 추출
# ─────────────────────────────────────────────────
def _extract_property_number(filename: str) -> str:
    """파일명 앞 7자리 숫자를 매물번호로 추출. '1234567_도면.jpg' → '1234567'"""
    if not filename:
        return ""
    m = re.match(r"^(\d{7})", filename.strip())
    if m:
        return m.group(1)
    return ""


# ──────────────────────────────────────────────────────
# 物件DB 탭 — 표시용 헬퍼 (4-1: 보기 전용)
# ──────────────────────────────────────────────────────
def _fmt_money(n) -> str:
    """숫자 → '134,000엔'. None/빈값이면 '-'."""
    try:
        if n is None or n == "":
            return "-"
        return f"{int(round(float(n))):,}엔"
    except (ValueError, TypeError):
        return str(n)


def _fmt_deposit_display(field) -> str:
    """
    시키킹/레이킹 {value, unit} → 표시 문자열 (원본값 그대로).
    - months: '1.0개월분'  / yen: '120,000엔'  / 값없음·판독불가: '확인필요'
    ※ 금액→개월수 환산(§4)은 제안서 생성 단계에서 적용. 여기선 원본만 표시.
    """
    if not isinstance(field, dict):
        return "확인필요"
    val, unit = field.get("value"), field.get("unit")
    if val is None or unit is None:
        return "확인필요"
    if unit == "months":
        try:
            return f"{float(val):.1f}개월분"
        except (ValueError, TypeError):
            return "확인필요"
    if unit == "yen":
        return _fmt_money(val)
    return "확인필요"


def _fmt_station_time(station) -> str:
    """역명(일/한)으로 신주쿠 소요시간 조회. 없거나 실패 시 ''."""
    if not station or sheet_sync is None:
        return ""
    try:
        return sheet_sync.get_station_time(station) or ""
    except Exception:
        return ""


def _render_property_card(p: dict) -> None:
    """物件DB 매물 카드 1장 렌더 (4-1: 보기 전용, 수정·선택 없음)."""
    data = p.get("data") or {}
    manual = p.get("manual_fields") or {}
    station_obj = data.get("nearest_station") or {}
    station = station_obj.get("station") or ""
    walk = station_obj.get("walk_minutes")

    with st.container(border=True):
        prop_num = p.get("property_number") or "-"
        bldg = data.get("property_name") or "(건물명 없음)"
        addr = data.get("address") or "-"
        st.markdown(
            f"**{bldg}**  \n"
            f"<span style='color:gray;font-size:0.85em'>{prop_num} · {addr}</span>",
            unsafe_allow_html=True,
        )
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(
                f"**월세/관리**  \n{_fmt_money(data.get('rent_yen'))} / "
                f"{_fmt_money(data.get('management_fee_yen'))}"
            )
            st.markdown(
                f"**시키킹/레이킹**  \n{_fmt_deposit_display(data.get('shikikin'))} / "
                f"{_fmt_deposit_display(data.get('reikin'))}"
            )
        with c2:
            area = data.get("area_sqm")
            st.markdown(
                f"**구조/면적**  \n{data.get('layout') or '-'} · "
                f"{area if area else '-'}㎡"
            )
            walk_txt = f"도보{walk}분" if walk else ""
            st.markdown(f"**가까운역**  \n{station or '-'} {walk_txt}")
        with c3:
            sinjuku = _fmt_station_time(station)
            st.markdown(f"**신주쿠**  \n{sinjuku or '-'}")
            photo = "입력됨" if manual.get("photo_link") else "미입력"
            st.markdown(f"**사진링크**  \n{photo}")


def _property_to_row(p: dict) -> dict:
    """物件DB 매물 1건 → 표(테이블) 한 행 dict. 한 줄에 모든 정보."""
    data = p.get("data") or {}
    manual = p.get("manual_fields") or {}
    st_obj = data.get("nearest_station") or {}
    station = st_obj.get("station") or ""
    walk = st_obj.get("walk_minutes")
    area = data.get("area_sqm")
    return {
        "매물번호": p.get("property_number") or "",
        "건물명": data.get("property_name") or "",
        "주소": data.get("address") or "",
        "월세": _fmt_money(data.get("rent_yen")),
        "관리비": _fmt_money(data.get("management_fee_yen")),
        "시키킹": _fmt_deposit_display(_effective_deposit(p, "shikikin")),
        "레이킹": _fmt_deposit_display(_effective_deposit(p, "reikin")),
        "구조": data.get("layout") or "",
        "면적": f"{area}㎡" if area else "",
        "가까운역": station,
        "도보": f"도보{walk}분" if walk else "",
        "신주쿠": _fmt_station_time(station) or "",
        "사진링크": "입력됨" if manual.get("photo_link") else "미입력",
    }


def _effective_deposit(p: dict, key: str):
    """시키킹/레이킹의 '유효값' — 직원 수정값(manual)이 있으면 그것, 없으면 추출값."""
    manual = p.get("manual_fields") or {}
    if isinstance(manual.get(key), dict):
        return manual[key]
    return (p.get("data") or {}).get(key)


# ──────────────────────────────────────────────────────
# 전체 칼럼 표 (4단계 전면 개편) — 한글 변환·행정구역·입주일·전체행
# ──────────────────────────────────────────────────────
import re as _re_fields
import urllib.parse as _urlparse


def _yn(v) -> str:
    """bool → 있음/없음/'' (None은 공란)."""
    if v is True:
        return "있음"
    if v is False:
        return "없음"
    return ""


def _ward_jp_from_address(addr: str) -> str:
    """주소에서 도도후켄(東京都/○○県 등) 떼고 행정구역부터. (일본어)"""
    if not addr:
        return ""
    return _re_fields.sub(r'^(東京都|北海道|(?:京都|大阪)府|.{2,3}県)', '', addr).strip()


def _move_in_no_year(s) -> str:
    """입주일에서 년도 제거. '2025年4月15日'→'4月15日', '即入居'→그대로."""
    if not s:
        return ""
    return _re_fields.sub(r'^\d{4}\s*[年/\-\.]\s*', '', str(s)).strip()


def _line_ko(jp: str) -> str:
    if not jp:
        return ""
    if translation_db is not None:
        try:
            return translation_db.translate_line(jp)
        except Exception:
            return jp
    return jp


def _station_ko(jp: str) -> str:
    if not jp:
        return ""
    if translation_db is not None:
        try:
            return translation_db.translate_station(jp)
        except Exception:
            return jp
    return jp


def _ward_ko(addr: str) -> str:
    if not addr:
        return ""
    if translation_db is not None:
        try:
            return translation_db.extract_korean_ward(addr)
        except Exception:
            return ""
    return ""


def _structure_ko(jp: str) -> str:
    if field_map is not None:
        try:
            return field_map.structure_ko(jp or "")
        except Exception:
            return jp or ""
    return jp or ""


def _direction_ko(jp: str) -> str:
    if field_map is not None:
        try:
            return field_map.direction_ko(jp or "")
        except Exception:
            return jp or ""
    return jp or ""


def _visa_text(filename: str) -> str:
    if field_map is not None:
        try:
            return field_map.visa_text(filename or "")
        except Exception:
            return ""
    return ""


def _deposit_months(field) -> float:
    """{value,unit} → 개월분 숫자. months면 그 값, 그 외(엔/없음)는 0.0."""
    if isinstance(field, dict) and field.get("unit") == "months":
        try:
            return float(field.get("value") or 0)
        except (ValueError, TypeError):
            return 0.0
    return 0.0


# 표 칼럼 순서 — 매물번호·입주일·비자를 앞으로, 도시가스 삭제. 전체 편집 가능.
_FULL_COLS = [
    "매물번호", "입주일", "비자",
    "건물명", "주소", "지역", "맵", "매물검색", "사진링크",
    "월세", "관리비", "월세+관리비", "시키킹", "레이킹", "방향",
    "노선1", "가까운역1", "도보1", "노선2", "가까운역2", "도보2", "신주쿠까지",
    "간취", "면적", "구조", "건물층수", "입주층", "건축연도",
    "인터넷", "엘리베이터", "택배박스", "오토록", "에어컨", "화장실욕실분리",
    "실내세탁", "독립세면대", "IH", "가스종류", "24시간쓰레기",
    "관리회사",
]
# 시키킹·레이킹=구조화 저장, 사진링크=photo_link, 그 외=overrides(텍스트)
_DEPOSIT_COLS = ["시키킹", "레이킹"]


def _property_to_full_row(p: dict) -> dict:
    """物件DB 매물 1건 → 전체 칼럼 표의 한 행 (한글 변환 적용)."""
    data = p.get("data") or {}
    manual = p.get("manual_fields") or {}
    fac = data.get("facilities") or {}
    cond = data.get("conditions") or {}
    ns = data.get("nearest_station") or {}
    adds = data.get("additional_stations") or []
    add1 = adds[0] if adds else {}

    bldg = data.get("property_name") or ""
    addr_jp = _ward_jp_from_address(data.get("address") or "")
    rent = data.get("rent_yen")
    mgmt = data.get("management_fee_yen")
    try:
        total = (int(rent) + int(mgmt)) if (rent is not None and mgmt is not None) else None
    except (ValueError, TypeError):
        total = None

    shiki_m = _deposit_months(_effective_deposit(p, "shikikin"))
    reiki_m = _deposit_months(_effective_deposit(p, "reikin"))

    station_jp = ns.get("station") or ""
    map_q = _urlparse.quote(f"{bldg} {data.get('address') or ''}".strip())
    search_q = _urlparse.quote(bldg) if bldg else ""

    gas_jp = fac.get("gas_type") or ""
    gas_ko = {"都市ガス": "도시가스", "プロパン": "프로판가스", "プロパンガス": "프로판가스"}.get(gas_jp, gas_jp or "")

    row = {
        "매물번호": p.get("property_number") or "",
        "건물명": bldg,
        "주소": addr_jp,
        "지역": _ward_ko(data.get("address") or ""),
        "맵": f"https://www.google.com/maps/search/{map_q}" if map_q else "",
        "매물검색": f"https://www.google.com/search?q={search_q}" if search_q else "",
        "사진링크": manual.get("photo_link") or "",
        "월세": _fmt_money(rent),
        "관리비": _fmt_money(mgmt),
        "월세+관리비": _fmt_money(total),
        "시키킹": shiki_m,
        "레이킹": reiki_m,
        "방향": _direction_ko(data.get("facing_direction") or ""),
        "노선1": _line_ko(ns.get("line") or ""),
        "가까운역1": _station_ko(station_jp),
        "도보1": f"도보{ns.get('walk_minutes')}분" if ns.get("walk_minutes") else "",
        "노선2": _line_ko(add1.get("line") or ""),
        "가까운역2": _station_ko(add1.get("station") or ""),
        "도보2": f"도보{add1.get('walk_minutes')}분" if add1.get("walk_minutes") else "",
        "신주쿠까지": _fmt_station_time(station_jp) or "",
        "간취": data.get("layout") or "",
        "면적": f"{data.get('area_sqm')}㎡" if data.get("area_sqm") else "",
        "구조": _structure_ko(data.get("structure") or ""),
        "건물층수": data.get("total_floors") or "",
        "입주층": data.get("floor") or "",
        "건축연도": data.get("construction_year") or "",
        "입주일": _move_in_no_year(data.get("available_from") or ""),
        "인터넷": ("무료" if fac.get("internet_free") is True else _yn(fac.get("internet_free"))),
        "엘리베이터": _yn(fac.get("elevator")),
        "택배박스": _yn(fac.get("delivery_box")),
        "오토록": _yn(fac.get("auto_lock")),
        "에어컨": _yn(fac.get("air_conditioner")),
        "화장실욕실분리": ("분리형" if fac.get("separate_bath_toilet") is True else _yn(fac.get("separate_bath_toilet"))),
        "실내세탁": ("실내" if fac.get("washing_machine_indoor") is True else _yn(fac.get("washing_machine_indoor"))),
        "독립세면대": _yn(fac.get("independent_washstand") if "independent_washstand" in fac else fac.get("separate_washstand")),
        "IH": _yn(None),  # 스키마에 직접 없음 → 공란(수동 입력으로 채움)
        "가스종류": gas_ko,
        "24시간쓰레기": "",  # 스키마에 직접 없음 → 수동 입력으로 채움
        "비자": _visa_text(p.get("filename") or ""),
        "관리회사": data.get("management_company") or "",
    }

    # 수동 수정값(overrides) 덮어쓰기 — 시키킹·레이킹·사진링크는 별도 저장이라 제외
    ov = manual.get("overrides") or {}
    for _k, _v in ov.items():
        if _k in row and _k not in ("시키킹", "레이킹", "사진링크"):
            row[_k] = _v
    return row


def _deposit_edit_widget(label: str, current, key: str, col):
    """
    시키킹/레이킹 수정 위젯. 반환: {"value","unit"} 또는 None(미입력/모름).
    current: 현재 유효값 {"value","unit"} (없으면 None)
    """
    units = ["모름/미입력", "개월분", "엔"]
    cur_label, cur_val = "모름/미입력", 0.0
    if isinstance(current, dict) and current.get("unit"):
        cur_label = {"months": "개월분", "yen": "엔"}.get(current.get("unit"), "모름/미입력")
        try:
            cur_val = float(current.get("value") or 0)
        except (ValueError, TypeError):
            cur_val = 0.0
    with col:
        st.markdown(f"**{label}**")
        unit_label = st.selectbox(
            "단위", units, index=units.index(cur_label),
            key=f"{key}_unit",
        )
        val = st.number_input(
            "값", min_value=0.0, value=cur_val, step=0.5,
            key=f"{key}_val", help="개월분이면 1.0, 0.5 등 / 엔이면 금액 입력",
        )
    if unit_label == "모름/미입력":
        return None
    return {"value": val, "unit": "months" if unit_label == "개월분" else "yen"}


def _render_property_editor(p: dict) -> None:
    """物件DB 매물 검토·수정 영역 (4-2: 사진링크 + 시키킹·레이킹 → 저장)."""
    import urllib.parse

    data = p.get("data") or {}
    manual = dict(p.get("manual_fields") or {})
    row_id = p.get("id")
    bldg = data.get("property_name") or ""

    with st.container(border=True):
        st.markdown(f"**{p.get('property_number') or ''} · {bldg or '(건물명 없음)'}**")

        # 사진 링크 (수동 입력) + 건물명 구글검색
        cur_photo = manual.get("photo_link") or ""
        photo = st.text_input(
            "사진 링크 (URL 붙여넣기)", value=cur_photo,
            key=f"photo_{row_id}", placeholder="https://...",
        )
        if bldg:
            _gq = urllib.parse.quote(bldg)
            st.markdown(f"🔍 [‘{bldg}’ 구글에서 검색](https://www.google.com/search?q={_gq})")

        # 시키킹 / 레이킹 수정
        c1, c2 = st.columns(2)
        new_shiki = _deposit_edit_widget("시키킹(敷金)", _effective_deposit(p, "shikikin"), f"shiki_{row_id}", c1)
        new_reiki = _deposit_edit_widget("레이킹(礼金)", _effective_deposit(p, "reikin"), f"reiki_{row_id}", c2)

        if st.button("💾 저장", key=f"save_{row_id}", type="primary"):
            if photo.strip():
                manual["photo_link"] = photo.strip()
            else:
                manual.pop("photo_link", None)
            if new_shiki is not None:
                manual["shikikin"] = new_shiki
            else:
                manual.pop("shikikin", None)
            if new_reiki is not None:
                manual["reikin"] = new_reiki
            else:
                manual.pop("reikin", None)
            try:
                property_db.update_manual_fields(row_id, manual)
                st.success("저장되었습니다.")
                st.rerun()
            except Exception as e:
                st.error(f"저장 실패: {e}")


def _insert_property_number_to_table(html: str, prop_num: str) -> str:
    """본문 '매물 기본정보' 표 맨 위에 매물번호 행을 삽입.
    첫 번째 <table>의 첫 <tr> 앞에 매물번호 행 추가.
    이미 매물번호 행이 있으면 중복 삽입 안 함.
    """
    if not html or not prop_num:
        return html
    # 이미 매물번호가 들어있으면 스킵
    if "매물번호" in html:
        return html

    # 첫 번째 <table ...> 다음에 매물번호 행 삽입
    num_row = (
        '<tr>'
        '<td style="border:1px solid #ddd;padding:8px 12px;background:#f5f5f5;'
        'width:30%;font-weight:bold">매물번호</td>'
        f'<td style="border:1px solid #ddd;padding:8px 12px">{prop_num}</td>'
        '</tr>'
    )
    # <table ...> 태그를 찾아 그 직후에 삽입
    m = re.search(r"(<table[^>]*>)", html)
    if m:
        insert_pos = m.end()
        return html[:insert_pos] + num_row + html[insert_pos:]
    # 표가 없으면 원본 유지
    return html


def _extract_ward(title: str) -> str:
    """제목에서 구/시 추출. 실제 제목: '이타바시구 도부토조선 나리마스역 도보 5분 ...'
    또는 매물번호가 앞에 붙은 경우: '[1234567] 이타바시구 ...' """
    if not title:
        return "기타"
    # 매물번호 [숫자] 제거
    t = re.sub(r"^\[\d+\]\s*", "", title.strip())
    # 첫 단어가 '~구' 또는 '~시'로 끝나면 그것이 지역
    m = re.match(r"^([가-힣]+(?:구|시|정|초|쿠))", t)
    if m:
        return m.group(1)
    # fallback: 첫 단어
    first_word = t.split()[0] if t.split() else ""
    return first_word if first_word else "기타"


def _extract_line(title: str) -> str:
    """제목에서 노선 추출. 실제 제목: '이타바시구 도부토조선 나리마스역 ...'
    구 다음 ~선/~라인으로 끝나는 단어."""
    if not title:
        return "기타"
    t = re.sub(r"^\[\d+\]\s*", "", title.strip())
    # ~선 또는 ~라인으로 끝나는 단어 찾기 (역 앞)
    m = re.search(r"([\w가-힣]+(?:선|라인|Line))\s+[\w가-힣]+역", t)
    if m:
        return m.group(1)
    # JR 야마노테선 같은 복합 노선
    m2 = re.search(r"(JR\s*[\w가-힣]+선)", t)
    if m2:
        return m2.group(1).replace(" ", "")
    return "기타"


def _extract_station(title: str) -> str:
    """제목에서 역명 추출. 실제 제목: '... 나리마스역 도보 5분 ...'
    ~역으로 끝나는 단어 (역 글자 제외)."""
    if not title:
        return "기타"
    t = re.sub(r"^\[\d+\]\s*", "", title.strip())
    # ~역 패턴 (역명만 추출)
    m = re.search(r"([\w가-힣]+)역", t)
    if m:
        return m.group(1)
    return "기타"


def _extract_rent(title: str) -> int:
    """제목에서 월세 추출 (엔 단위). '월세 ¥80,000+관리비 ¥5,000' → 85000"""
    if not title:
        return 0
    total = 0
    # 월세 ¥XX,XXX (공백 허용)
    m_rent = re.search(r"월세\s*[¥￥]\s*([\d,]+)", title)
    if m_rent:
        try:
            total += int(m_rent.group(1).replace(",", ""))
        except ValueError:
            pass
    # 관리비 ¥X,XXX (공백 허용)
    m_mgmt = re.search(r"관리비\s*[¥￥]\s*([\d,]+)", title)
    if m_mgmt:
        try:
            total += int(m_mgmt.group(1).replace(",", ""))
        except ValueError:
            pass
    return total


def _extract_room_type(title: str) -> str:
    """제목에서 방구조 추출. '도보 5분 1K 월세...' → '1K'"""
    if not title:
        return "기타"
    # 도보 N분 [방구조] 월세 패턴 (공백 허용)
    m = re.search(r"도보\s*\d+\s*분\s+([\w\d]+)\s+월세", title)
    if m:
        return m.group(1)
    return "기타"


def _aggregate_by_month(history: list) -> dict:
    """월별 매물 건수 집계: {'2026-05': 12, '2026-06': 8, ...}"""
    counter = {}
    for h in history:
        ts = h.get("timestamp", "")
        if ts:
            month = ts[:7]  # 'YYYY-MM'
            counter[month] = counter.get(month, 0) + 1
    # 시간순 정렬
    return dict(sorted(counter.items()))


def _aggregate_by_day(history: list, days: int = 30) -> dict:
    """최근 N일간 일별 매물 건수: {'2026-05-27': 3, '2026-05-26': 2, ...}"""
    counter = {}
    cutoff = datetime.now() - timedelta(days=days)
    cutoff_iso = cutoff.isoformat(timespec="seconds")
    for h in history:
        ts = h.get("timestamp", "")
        if ts and ts >= cutoff_iso:
            day = ts[:10]  # 'YYYY-MM-DD'
            counter[day] = counter.get(day, 0) + 1
    # 시간순 정렬
    return dict(sorted(counter.items()))


def _aggregate_by_ward(history: list) -> dict:
    """구별 매물 건수: {'이타바시구': 8, '신주쿠구': 5, ...} - 내림차순"""
    counter = {}
    for h in history:
        ward = _extract_ward(h.get("title", ""))
        counter[ward] = counter.get(ward, 0) + 1
    return dict(sorted(counter.items(), key=lambda x: -x[1]))


def _aggregate_by_station(history: list, top_n: int = 10) -> dict:
    """역별 매물 건수 TOP N: {'이케부쿠로': 3, '신주쿠': 2, ...} - 내림차순"""
    counter = {}
    for h in history:
        station = _extract_station(h.get("title", ""))
        counter[station] = counter.get(station, 0) + 1
    sorted_items = sorted(counter.items(), key=lambda x: -x[1])
    return dict(sorted_items[:top_n])


def _aggregate_by_line(history: list, top_n: int = 10) -> dict:
    """노선별 매물 건수 TOP N - 내림차순"""
    counter = {}
    for h in history:
        line = _extract_line(h.get("title", ""))
        counter[line] = counter.get(line, 0) + 1
    sorted_items = sorted(counter.items(), key=lambda x: -x[1])
    return dict(sorted_items[:top_n])


def _calc_rent_stats(history: list) -> dict:
    """월세 통계: 평균·최저·최고·중앙값"""
    rents = []
    for h in history:
        rent = _extract_rent(h.get("title", ""))
        if rent > 0:
            rents.append(rent)
    if not rents:
        return {"count": 0, "avg": 0, "min": 0, "max": 0, "median": 0}
    rents_sorted = sorted(rents)
    n = len(rents_sorted)
    median = rents_sorted[n // 2] if n % 2 == 1 else (rents_sorted[n // 2 - 1] + rents_sorted[n // 2]) // 2
    return {
        "count": n,
        "avg": sum(rents) // n,
        "min": min(rents),
        "max": max(rents),
        "median": median,
    }


def _format_yen(amount: int) -> str:
    """엔 금액을 보기 좋게 포맷: 85000 → '¥85,000'"""
    if amount <= 0:
        return "¥0"
    return f"¥{amount:,}"


def _build_excel_report(history: list) -> bytes:
    """이력 전체를 Excel(xlsx)로 변환. openpyxl 사용."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return b""

    wb = Workbook()
    ws = wb.active
    ws.title = "매물 이력"

    # 헤더
    headers = ["번호", "제목", "구", "노선", "역", "방구조", "월세+관리비(엔)", "파일명", "생성일시", "즐겨찾기", "해시태그"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    for idx, h in enumerate(history, start=1):
        title = h.get("title", "")
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=title)
        ws.cell(row=idx + 1, column=3, value=_extract_ward(title))
        ws.cell(row=idx + 1, column=4, value=_extract_line(title))
        ws.cell(row=idx + 1, column=5, value=_extract_station(title))
        ws.cell(row=idx + 1, column=6, value=_extract_room_type(title))
        ws.cell(row=idx + 1, column=7, value=_extract_rent(title))
        ws.cell(row=idx + 1, column=8, value=h.get("filename", ""))
        ws.cell(row=idx + 1, column=9, value=h.get("timestamp", ""))
        ws.cell(row=idx + 1, column=10, value="⭐" if h.get("favorite") else "")
        tags = h.get("hashtags", [])
        ws.cell(row=idx + 1, column=11, value=" ".join(tags) if tags else "")

    # 컬럼 너비 조정
    widths = [6, 50, 12, 18, 14, 10, 16, 22, 18, 8, 30]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    # 통계 시트
    ws2 = wb.create_sheet("통계 요약")
    ws2.cell(row=1, column=1, value="구별 매물 분포").font = Font(bold=True, size=12)
    by_ward = _aggregate_by_ward(history)
    for i, (k, v) in enumerate(by_ward.items(), start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)

    row_offset = len(by_ward) + 4
    ws2.cell(row=row_offset, column=1, value="역별 TOP 10").font = Font(bold=True, size=12)
    by_station = _aggregate_by_station(history, 10)
    for i, (k, v) in enumerate(by_station.items(), start=row_offset + 1):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10

    # 바이트로 변환
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────
# 인증 시스템 — 이메일 + 비밀번호 (역할 기반)
# ─────────────────────────────────────────────────
from src.persistence import (
    init_admin_if_needed,
    authenticate_user,
    add_user as db_add_user,
    delete_user as db_delete_user,
    change_password as db_change_password,
    list_users as db_list_users,
    load_admin_settings,
    save_admin_settings,
    is_valid_email,
    ADMIN_EMAIL,
    ALLOWED_DOMAIN,
)

# 첫 실행 시 관리자 계정 자동 생성 (Render 환경변수 ADMIN_INITIAL_PASSWORD 사용)
_admin_initial_pw = os.getenv("ADMIN_INITIAL_PASSWORD", "").strip()
if _admin_initial_pw:
    init_admin_if_needed(_admin_initial_pw)


@st.cache_data(ttl=60, show_spinner=False)
def _cached_load_users() -> dict:
    """
    load_users() 결과를 60초 캐싱 — 로그인 화면에서 중복 호출 방지.
    TTL 60초: 새 계정 추가 후 1분 이내에는 반영됨.
    """
    from src.persistence import load_users
    return load_users()


def _email_token(email: str) -> str:
    """이메일로부터 URL 토큰 생성 (계정 기억용)"""
    seed = (email + os.getenv("ADMIN_INITIAL_PASSWORD", "secret")).encode("utf-8")
    return hashlib.sha256(seed).hexdigest()[:24]


def _check_login() -> bool:
    """로그인 확인. 성공 시 True, 로그인 화면 표시 시 False."""
    # ⭐ 네이버 OAuth callback 자동 우회 (이미 로그인된 상태)
    has_oauth_params = (
        st.query_params.get("code")
        or st.query_params.get("state")
        or st.query_params.get("error")
    )
    if has_oauth_params and st.session_state.get("user"):
        return True

    # URL 토큰으로 자동 로그인 (세션 무제한)
    url_email = st.query_params.get("user_email", "")
    url_token = st.query_params.get("auth", "")
    if url_email and url_token:
        expected = _email_token(url_email)
        if url_token == expected:
            # 토큰 검증되면 사용자 정보 다시 로드 (역할 변경 반영)
            users = _cached_load_users()
            info = users.get(url_email)
            if info:
                st.session_state["user"] = {
                    "email": url_email,
                    "role": info.get("role", "user"),
                }
                return True

    # session_state 직접 확인
    if st.session_state.get("user"):
        return True

    # 로그인 화면
    _render_login_screen()
    return False


def _render_login_screen():
    """로그인 화면 렌더링 — 가독성 좋은 디자인"""
    # 페이지 중앙 정렬
    _, col_center, _ = st.columns([1, 2, 1])
    with col_center:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(
            "<div style='text-align:center;'>"
            "<h1 style='margin-bottom:0;'>🏠 JRE일본부동산</h1>"
            "<p style='color:#666;margin-top:0;'>네이버 블로그 자동작성 시스템</p>"
            "</div>",
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

        # 관리자 계정 미설정 시 안내
        users = _cached_load_users()
        if not users:
            st.error(
                "⚠️ **관리자 계정이 아직 설정되지 않았습니다.**\n\n"
                "Render Dashboard → Environment에서 "
                "`ADMIN_INITIAL_PASSWORD` 환경변수를 추가한 후 재배포해주세요."
            )
            st.stop()

        with st.container(border=True):
            st.markdown("### 🔐 로그인")

            with st.form("login_form_v2"):
                # 이메일 기억 기능 — 최근 로그인 이메일 자동 입력
                remembered_email = st.session_state.get("remembered_email", "")
                email = st.text_input(
                    "📧 이메일",
                    value=remembered_email,
                    placeholder="info@win-bro.com",
                    autocomplete="email",
                )
                password = st.text_input(
                    "🔒 비밀번호",
                    type="password",
                    autocomplete="current-password",
                )
                submitted = st.form_submit_button("로그인", type="primary", use_container_width=True)

                if submitted:
                    user = authenticate_user(email, password)
                    if user:
                        # 세션에 저장
                        st.session_state["user"] = {
                            "email": user["email"],
                            "role": user["role"],
                        }
                        st.session_state["remembered_email"] = user["email"]
                        # URL 토큰에 이메일 정보 저장 (계정 기억)
                        st.query_params["user_email"] = user["email"]
                        st.query_params["auth"] = _email_token(user["email"])
                        # ⭐ 로그인 시 새 세션 ID 발급 + 이전 작업 상태 초기화
                        # (계정 전환 시 이전 계정의 분석 결과가 따라오지 않게)
                        st.query_params["sid"] = generate_session_id()
                        st.session_state.pop("properties", None)
                        st.session_state.pop("blog_posts", None)
                        st.session_state.pop("untranslated_alert", None)
                        st.session_state.pop("_analysis_done_this_session", None)
                        st.session_state.pop("_session_restored", None)
                        st.rerun()
                    else:
                        st.error("❌ 이메일 또는 비밀번호가 올바르지 않습니다.")

        st.markdown(
            "<br><div style='text-align:center;color:#999;font-size:12px;'>"
            "© WinBro LLC · JRE일본부동산"
            "</div>",
            unsafe_allow_html=True,
        )


# 로그인 확인 — 통과 못 하면 여기서 멈춤
if not _check_login():
    st.stop()


# 로그인된 사용자 정보 (전역에서 사용)
current_user = st.session_state.get("user", {})
current_email = current_user.get("email", "")
current_role = current_user.get("role", "user")
is_admin = (current_role == "admin")


def _logout():
    """로그아웃 — 세션 + URL 토큰 + 작업 상태 모두 제거"""
    st.session_state.pop("user", None)
    # remembered_email은 유지 (다음 로그인 시 자동 입력)
    if "user_email" in st.query_params:
        del st.query_params["user_email"]
    if "auth" in st.query_params:
        del st.query_params["auth"]
    # ⭐ 작업 상태 + 세션 ID 초기화 (다음 로그인 계정에 안 따라가게)
    st.session_state.pop("properties", None)
    st.session_state.pop("blog_posts", None)
    st.session_state.pop("untranslated_alert", None)
    st.session_state.pop("_analysis_done_this_session", None)
    st.session_state.pop("_session_restored", None)
    if "sid" in st.query_params:
        del st.query_params["sid"]





# ─────────────────────────────────────────────────
# 세션 ID 관리 + 자동 복원 (새로고침 시 작업 유지)
# ─────────────────────────────────────────────────
def _ensure_session_id() -> str:
    """
    URL 쿼리에 session ID 보장.
    - 이미 있으면 그 ID 사용 (새로고침해도 같음)
    - 없으면 새로 생성하고 URL에 추가
    """
    sid = st.query_params.get("sid", "")
    if not sid:
        sid = generate_session_id()
        st.query_params["sid"] = sid
    return sid


def _restore_session_if_needed():
    """
    페이지 첫 로드 시 디스크에서 세션 데이터 복원.
    session_state에 이미 데이터가 있으면 (= 같은 탭 진행 중) 복원 안 함.
    """
    if st.session_state.get("_session_restored"):
        return  # 이미 복원 시도함

    sid = st.query_params.get("sid", "")
    if not sid:
        return

    # 분석/생성을 한 번이라도 한 세션이면 복원하지 않음
    # (새 분석 결과를 디스크의 옛 데이터가 덮어쓰는 버그 방지)
    if st.session_state.get("_analysis_done_this_session"):
        st.session_state["_session_restored"] = True
        return

    # 새로고침 직후엔 session_state가 비어 있음
    # 디스크에서 복원 시도
    saved = load_session(sid)
    if saved:
        if saved.get("properties") and not st.session_state.get("properties"):
            st.session_state["properties"] = saved["properties"]
        if saved.get("blog_posts") and not st.session_state.get("blog_posts"):
            st.session_state["blog_posts"] = saved["blog_posts"]

    st.session_state["_session_restored"] = True


def _persist_session(overwrite: bool = False):
    """현재 작업 상태를 디스크에 자동 저장.
    overwrite=True면 properties가 없어도 빈 값으로 덮어써서
    옛 데이터가 남지 않게 함.
    """
    sid = st.query_params.get("sid", "")
    if not sid:
        return

    data = {}
    if st.session_state.get("properties"):
        data["properties"] = st.session_state["properties"]
    if st.session_state.get("blog_posts"):
        data["blog_posts"] = st.session_state["blog_posts"]

    if data or overwrite:
        save_session(sid, data)


# 페이지 첫 로드 시: 세션 ID 보장 + 자동 복원 + 오래된 세션 정리
_ensure_session_id()
_restore_session_if_needed()

# 임시 세션 파일 자동 정리 (24시간 이상 된 것)
if not st.session_state.get("_cleanup_done"):
    cleanup_old_sessions()
    st.session_state["_cleanup_done"] = True


st.markdown(
    '<a href="javascript:void(0)" onclick="window.location.reload();" '
    'style="text-decoration:none;color:inherit;cursor:pointer;display:inline-block;" '
    'title="클릭하면 새로고침">'
    '<h1 style="margin:0 0 0.25rem 0;padding:0;">🏠 JRE일본부동산</h1>'
    '</a>',
    unsafe_allow_html=True,
)
st.caption(
    f"네이버 블로그 자동작성 시스템 · 마이소크 최대 {MAX_UPLOADS}개 업로드 → "
    "AI 병렬 분석 → 한국어 블로그 일괄 생성"
)

# 사용자 환영 카드 — 회사 전체 통계 표시 (이력은 모두 공유)
try:
    _all_history = load_history()
    # 기존 이력 호환성 (user_email 없는 경우)
    for h in _all_history:
        if not h.get("user_email"):
            h["user_email"] = ADMIN_EMAIL

    # 회사 전체 매물 (관리자/사용자 모두 동일하게 표시)
    _shown_history = _all_history

    _my_count = len(_shown_history)
    _my_cost = _my_count * 110  # 매물당 약 110원 (Opus + Extended Thinking)
    _my_this_month = sum(
        1 for h in _shown_history
        if h.get("timestamp", "")[:7] == datetime.now().strftime("%Y-%m")
    )
except Exception:
    _my_count = _my_cost = _my_this_month = 0

# 라벨 — 역할만 다르고 통계는 모두 전체
_role_label = "👑 관리자" if is_admin else "👤 사용자"
_count_label = "회사 매물"
st.markdown(
    f"""
<div style='background:linear-gradient(135deg,#E3F2FD 0%,#F3E5F5 100%);
            padding:14px 18px;border-radius:10px;margin-bottom:12px;
            border:1px solid #BBDEFB;'>
  <div style='display:flex;justify-content:space-between;align-items:center;
              flex-wrap:wrap;gap:12px;'>
    <div>
      <span style='font-size:11px;color:#1976D2;font-weight:600;'>{_role_label}</span><br>
      <span style='font-size:15px;color:#333;font-weight:600;'>{current_email}</span>
    </div>
    <div style='display:flex;gap:20px;flex-wrap:wrap;'>
      <div style='text-align:center;'>
        <div style='font-size:11px;color:#666;'>{_count_label}</div>
        <div style='font-size:18px;font-weight:700;color:#1976D2;'>{_my_count}건</div>
      </div>
      <div style='text-align:center;'>
        <div style='font-size:11px;color:#666;'>이번 달</div>
        <div style='font-size:18px;font-weight:700;color:#2E7D32;'>{_my_this_month}건</div>
      </div>
      <div style='text-align:center;'>
        <div style='font-size:11px;color:#666;'>AI 비용</div>
        <div style='font-size:18px;font-weight:700;color:#F57C00;'>{_my_cost:,}원</div>
      </div>
    </div>
  </div>
</div>
""",
    unsafe_allow_html=True,
)


# ───── 통계 표시 함수 (사이드바에서 호출) ─────
def _render_company_stats():
    """회사 전체 통계 - 모든 사용자 공개"""
    try:
        history_all = load_history()
        for h in history_all:
            if not h.get("user_email"):
                h["user_email"] = ADMIN_EMAIL
    except Exception:
        history_all = []

    if not history_all:
        st.info("아직 매물 데이터가 없습니다.")
        return

    # 4개 요약 카드
    cc1, cc2 = st.columns(2)
    with cc1:
        st.metric("📊 전체", f"{len(history_all)}건")
    with cc2:
        active_users = set(h.get("user_email", ADMIN_EMAIL) for h in history_all)
        st.metric("👥 사용자", f"{len(active_users)}명")

    cc3, cc4 = st.columns(2)
    with cc3:
        this_month = datetime.now().strftime("%Y-%m")
        this_month_count = sum(
            1 for h in history_all
            if h.get("timestamp", "")[:7] == this_month
        )
        st.metric("📅 이번달", f"{this_month_count}건")
    with cc4:
        fav_count = sum(1 for h in history_all if h.get("favorite"))
        st.metric("⭐ 즐겨찾기", f"{fav_count}건")

    st.markdown("**📅 월별 추이**")
    by_month = _aggregate_by_month(history_all)
    if by_month:
        st.bar_chart(by_month, height=150)

    # 월세 통계
    rent_stats = _calc_rent_stats(history_all)
    if rent_stats["count"] > 0:
        st.markdown("**💰 월세 통계**")
        st.caption(
            f"평균: {_format_yen(rent_stats['avg'])} · "
            f"중앙값: {_format_yen(rent_stats['median'])}"
        )
        st.caption(
            f"최저: {_format_yen(rent_stats['min'])} · "
            f"최고: {_format_yen(rent_stats['max'])}"
        )

    # 분포 (사이드바 좁아서 간소화)
    st.markdown("**🏷 구별 TOP 5**")
    by_ward = _aggregate_by_ward(history_all)
    if by_ward:
        for ward, cnt in list(by_ward.items())[:5]:
            st.caption(f"`{ward}` — **{cnt}건**")

    st.markdown("**🚉 역 TOP 5**")
    by_station = _aggregate_by_station(history_all, 5)
    if by_station:
        for station, cnt in by_station.items():
            st.caption(f"`{station}` — **{cnt}건**")


def _render_staff_stats():
    """직원별 작업 통계 - 모든 사용자 공개"""
    try:
        history_all = load_history()
        for h in history_all:
            if not h.get("user_email"):
                h["user_email"] = ADMIN_EMAIL
    except Exception:
        history_all = []

    if not history_all:
        st.info("아직 매물 데이터가 없습니다.")
        return

    st.caption("💡 각 직원의 작업량과 누적 AI 비용")

    # 사용자별 집계
    users_list = db_list_users()
    user_emails_set = {u["email"] for u in users_list}

    COST_PER_PROPERTY = 110

    user_stats = {}
    for h in history_all:
        uemail = h.get("user_email", ADMIN_EMAIL)
        if uemail not in user_stats:
            user_stats[uemail] = {
                "count": 0,
                "rent_total": 0,
                "rent_count": 0,
                "last": h.get("timestamp", ""),
            }
        user_stats[uemail]["count"] += 1
        rent = _extract_rent(h.get("title", ""))
        if rent > 0:
            user_stats[uemail]["rent_total"] += rent
            user_stats[uemail]["rent_count"] += 1
        ts = h.get("timestamp", "")
        if ts and ts > user_stats[uemail]["last"]:
            user_stats[uemail]["last"] = ts

    # 등록 사용자 + 매물 작성 사용자 모두 표시
    all_emails = user_emails_set | set(user_stats.keys())

    # 정렬: admin 먼저, 다음에 작업량 많은 순
    sorted_emails = sorted(all_emails, key=lambda e: (
        e != ADMIN_EMAIL,
        -user_stats.get(e, {}).get("count", 0),
        e,
    ))

    for uemail in sorted_emails:
        stats = user_stats.get(uemail, {"count": 0, "rent_total": 0, "rent_count": 0, "last": ""})
        u_info = next((u for u in users_list if u["email"] == uemail), None)
        u_role = u_info["role"] if u_info else "(삭제)"
        role_icon = "👑" if u_role == "admin" else "👤" if u_role == "user" else "❌"

        count = stats["count"]
        cost = count * COST_PER_PROPERTY

        # 이메일 짧게 표시 (사이드바 좁음)
        email_short = uemail.split("@")[0]

        with st.container(border=True):
            st.markdown(
                f"{role_icon} **{email_short}**  \n"
                f"<span style='color:#888;font-size:11px;'>{uemail}</span>",
                unsafe_allow_html=True,
            )
            sc1, sc2 = st.columns(2)
            sc1.metric("📝 매물", f"{count}건", label_visibility="visible")
            sc2.metric("💰 비용", f"{cost:,}원", label_visibility="visible")

    # 합계
    st.divider()
    total_props = sum(s["count"] for s in user_stats.values())
    total_cost = total_props * COST_PER_PROPERTY
    st.markdown(
        f"**📊 전사 합계**: {total_props}건 · "
        f"**💰 {total_cost:,}원**"
    )



# ─────────────────────────────────────────────────
# 사이드바
# ─────────────────────────────────────────────────
with st.sidebar:
    # ─── 사용자 정보 카드 (가독성 개선) ───
    role_badge = "👑 관리자" if is_admin else "👤 사용자"
    st.markdown(
        f"<div style='padding:12px;background:#F0F7FF;border-radius:8px;border:1px solid #BBDEFB;'>"
        f"<div style='font-size:11px;color:#1976D2;font-weight:600;margin-bottom:4px;'>{role_badge}</div>"
        f"<div style='font-size:13px;color:#333;font-weight:500;word-break:break-all;'>{current_email}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    if st.button("🚪 로그아웃", use_container_width=True, key="logout_btn"):
        _logout()
        st.rerun()

    st.divider()

    # ─── 작업 설정 (모든 사용자) ───
    st.markdown("### ⚙️ 작업 설정")

    target_visa = st.selectbox(
        "타깃 비자 (참고용)",
        options=list(VISA_LABELS.keys()),
        format_func=lambda k: VISA_LABELS[k],
        index=0,
        help="추천 이유는 비자별로 나누지 않고 통합 작성됩니다.",
    )

    available_styles = list_available_styles()
    default_style = st.selectbox(
        "📝 기본 글 스타일",
        options=available_styles,
        index=0,
        help="2번 탭에서 매물마다 다른 스타일로 변경할 수 있습니다.",
    )

    st.divider()

    # ─── 관리자 설정 (분석 엔진/AI 모델) — 일반 사용자는 admin이 저장한 기본값 자동 사용 ───
    _admin_settings = load_admin_settings()

    # 일반 사용자: admin이 설정한 값 자동 사용 (UI에 표시 안 함)
    if not is_admin:
        engine = _admin_settings.get("engine", "hybrid")
        model = _admin_settings.get("model", "claude-opus-4-7")

    # ─── 🗑 오래된 이력 일괄 정리 (관리자 전용 — 우선 표시) ───
    if is_admin:
        with st.expander("🗑 오래된 이력 일괄 정리 (관리자)", expanded=False):
            st.caption("💡 즐겨찾기(⭐)는 삭제되지 않습니다.")
            months_to_clean = st.slider(
                "몇 개월 이상 삭제?",
                min_value=1, max_value=24, value=12, step=1,
                key="cleanup_months_slider_sidebar",
            )
            confirm_key = f"_confirm_cleanup_sb_{months_to_clean}"
            if st.button(
                f"🗑 {months_to_clean}개월+ 삭제",
                type="secondary",
                use_container_width=True,
                key=f"cleanup_btn_sb_{months_to_clean}",
            ):
                if st.session_state.get(confirm_key):
                    deleted = delete_old_history(months_to_clean)
                    st.session_state.pop(confirm_key, None)
                    st.success(f"✅ {deleted}건 삭제 완료 (즐겨찾기 보호)")
                    st.rerun()
                else:
                    st.session_state[confirm_key] = True
                    st.warning("⚠️ 한 번 더 클릭하면 삭제됩니다.")

    # ─── 📥 Excel 리포트 다운로드 (모든 사용자) ───
    with st.expander("📥 Excel 리포트 다운로드", expanded=False):
        st.caption("💡 전체 매물 이력 + 통계 요약 Excel")
        try:
            _sb_history = load_history()
            for _h in _sb_history:
                if not _h.get("user_email"):
                    _h["user_email"] = ADMIN_EMAIL
            excel_bytes = _build_excel_report(_sb_history)
            if excel_bytes:
                st.download_button(
                    f"📊 다운로드 ({len(_sb_history)}건)",
                    excel_bytes,
                    file_name=f"JRE_매물이력_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            else:
                st.warning("⚠️ openpyxl 미설치")
        except Exception as e:
            st.error(f"Excel 생성 실패: {e}")

    # ─── 관리자 전용 영역 (사용자 관리 → 분석 엔진 → AI 모델 → 구글 연결 테스트) ───
    if is_admin:
        # ─── 사용자 관리 ───
        with st.expander("👥 사용자 관리 (관리자 전용)", expanded=False):
            users = db_list_users()
            st.caption(f"등록된 사용자: **{len(users)}명**")

            # 사용자 목록 표시
            for u in users:
                u_role = u["role"]
                u_email = u["email"]
                role_icon = "👑" if u_role == "admin" else "👤"
                last_login = u.get("last_login", "")
                last_login_str = last_login[:10] if last_login else "없음"

                with st.container(border=True):
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        st.markdown(
                            f"{role_icon} **{u_email}**  \n"
                            f"<span style='color:#888;font-size:11px;'>"
                            f"마지막 로그인: {last_login_str}</span>",
                            unsafe_allow_html=True,
                        )
                    with col2:
                        if u_role != "admin":
                            if st.button("🗑", key=f"del_user_{u_email}", help="이 사용자 삭제"):
                                ok, msg = db_delete_user(u_email)
                                if ok:
                                    st.success(msg)
                                    st.rerun()
                                else:
                                    st.error(msg)

            st.divider()
            st.markdown("**➕ 새 사용자 추가**")
            with st.form("add_user_form"):
                new_email = st.text_input(
                    "이메일",
                    placeholder=f"staff1{ALLOWED_DOMAIN}",
                    help=f"{ALLOWED_DOMAIN} 도메인만 허용",
                )
                new_password = st.text_input(
                    "비밀번호 (최소 4자)",
                    type="password",
                    help="이 비밀번호를 직원에게 직접 전달하세요.",
                )
                add_submitted = st.form_submit_button(
                    "➕ 사용자 추가",
                    type="primary",
                    use_container_width=True,
                )
                if add_submitted:
                    ok, msg = db_add_user(new_email, new_password, current_email)
                    if ok:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)

            st.divider()
            st.markdown("**🔒 비밀번호 변경**")
            with st.form("change_pw_form"):
                pw_email = st.selectbox(
                    "사용자",
                    options=[u["email"] for u in users],
                )
                new_pw = st.text_input(
                    "새 비밀번호",
                    type="password",
                )
                pw_submitted = st.form_submit_button(
                    "🔒 비밀번호 변경",
                    use_container_width=True,
                )
                if pw_submitted:
                    ok, msg = db_change_password(pw_email, new_pw)
                    if ok:
                        st.success(msg)
                    else:
                        st.error(msg)

        # ─── 🔬 분석 엔진 (관리자 전용) ───
        st.markdown("### 🔬 분석 엔진 (관리자 전용)")
        has_gemini = bool(os.getenv("GEMINI_API_KEY", "").strip())
        engine_options = {
            "hybrid": "🔀 하이브리드 (무료+Claude) ⭐ 추천",
            "gemini": "🆓 Gemini 무료만",
            "claude": "💎 Claude 유료만 (최고 정확도)",
        }
        engine_keys = list(engine_options.keys())
        try:
            engine_idx = engine_keys.index(_admin_settings.get("engine", "hybrid"))
        except ValueError:
            engine_idx = 0
        engine = st.selectbox(
            "분석 엔진 선택",
            options=engine_keys,
            format_func=lambda k: engine_options[k],
            index=engine_idx,
            help="이 설정은 모든 사용자에게 적용됩니다.",
        )
        if engine in ("hybrid", "gemini") and not has_gemini:
            st.warning(
                "⚠️ GEMINI_API_KEY가 설정되지 않았습니다. "
                "https://aistudio.google.com/apikey 에서 무료 발급 후 "
                "환경변수에 추가하세요. 현재는 Claude만 사용됩니다."
            )

        # ─── 🤖 AI 모델 (관리자 전용) ───
        st.markdown("### 🤖 AI 모델 (관리자 전용)")
        model_options = ["claude-opus-4-7", "claude-sonnet-4-6", "claude-haiku-4-5-20251001"]
        try:
            model_idx = model_options.index(_admin_settings.get("model", "claude-opus-4-7"))
        except ValueError:
            model_idx = 0
        model = st.selectbox(
            "Claude 모델",
            model_options,
            index=model_idx,
            help=(
                "Opus 4.7: ⭐ 권장 — 최고 정확도 (도면 분석·블로그 생성에 가장 정확)\n"
                "Sonnet 4.6: 균형형 — 정확도 양호 + 빠름\n"
                "Haiku 4.5: 가장 빠름 — 간단한 매물·테스트용\n\n"
                "이 설정은 모든 사용자에게 적용됩니다."
            ),
        )

        # 변경되면 즉시 저장 (모든 사용자에게 반영)
        if engine != _admin_settings.get("engine") or model != _admin_settings.get("model"):
            save_admin_settings(engine, model)
            st.caption(f"✅ 모든 사용자에게 적용됨")

        # ─── 🔗 Google Drive 연결 테스트 (관리자 전용 — 맨 마지막) ───
        with st.expander("🔗 Google Drive 연결 테스트 (관리자)", expanded=False):
            st.caption(
                "Phase 1(Google Cloud + Drive 설정)이 제대로 됐는지 확인하는 도구. "
                "관리자만 보임. 자동화 정상 작동 전 필수 검증."
            )
            if st.button("🔍 Drive 연결 확인", key="drive_verify_btn", use_container_width=True):
                with st.spinner("Drive API 호출 중…"):
                    try:
                        from src.drive_sync import verify_connection
                        result = verify_connection()
                    except ImportError as ie:
                        st.error(
                            f"⚠️ drive_sync 모듈 로드 실패: {ie}\n\n"
                            "→ `requirements.txt`에 `google-api-python-client`와 "
                            "`google-auth`를 추가하고 Render를 재배포했는지 확인하세요."
                        )
                        result = None

                if result is not None:
                    if result.get("ok"):
                        st.success("✅ 모든 검증 통과! Phase 3로 진행 가능합니다.")
                        st.markdown(
                            f"- **서비스 계정**: `{result['service_account_email']}`\n"
                            f"- **루트 폴더**: **{result['root_folder_name']}**  \n"
                            f"  (ID: `{result['root_folder_id']}`)\n"
                            f"- **쓰기 권한**: ✅ 편집자"
                        )
                    else:
                        st.error(f"❌ 검증 실패\n\n{result.get('error', '알 수 없는 오류')}")
                        with st.expander("진단 정보 (디버깅용)", expanded=False):
                            st.json({
                                k: v for k, v in result.items()
                                if k != "error"
                            })

    st.divider()
    st.caption(
        "💡 **사용 안내**\n\n"
        "여러 직원이 동시에 접속해도 각자의 작업이 분리됩니다.\n\n"
        "⚠️ 같은 시각에 분석을 동시 실행하면 메모리 부족 에러가 날 수 있으니, "
        "가능하면 5~10분 간격을 두고 사용하세요."
    )

    st.divider()
    st.caption(
        f"💾 작업 완료 후 ZIP을 받으시면 다음 폴더에 저장하시는 것을 권장합니다:\n\n"
        f"`{OUTPUT_FOLDER_PATH}`"
    )


# ─────────────────────────────────────────────────
# 병렬 분석 워커
# ─────────────────────────────────────────────────
def _analyze_worker(file_bytes: bytes, suffix: str, engine: str, model: str) -> dict:
    """ThreadPoolExecutor 워커: 임시 파일에 저장 후 분석, 자동 삭제."""
    with NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        return analyze_property_sheet(tmp_path, engine=engine, model=model)
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def _generate_worker(property_data, target_visa, style_name, custom_instructions, model):
    """ThreadPoolExecutor 워커: 블로그 생성."""
    return generate_blog_post(
        property_data=property_data,
        target_visa=target_visa,
        style_name=style_name,
        custom_instructions=custom_instructions,
        model=model,
    )


# ─────────────────────────────────────────────────
# 블로그 일괄 생성 공통 헬퍼 (Tab 1 자동 + Tab 2 수동 공통 사용)
# ─────────────────────────────────────────────────
def _run_blog_generation(
    properties: list,
    custom_instructions: str,
    target_visa: str,
    model: str,
    current_email: str,
) -> None:
    """
    블로그 일괄 생성 (병렬 처리) + 이력 영구 저장.
    - Tab 1 자동 모드: 분석 직후 즉시 호출
    - Tab 2 수동 모드: 사장님이 버튼 클릭 시 호출
    UI(progress·success·error)를 직접 렌더링하고, 완료 시 st.rerun() 호출.
    """
    blog_posts = [None] * len(properties)
    errors = []
    progress = st.progress(0.0, text="병렬 생성 시작…")

    with ThreadPoolExecutor(max_workers=MAX_PARALLEL_WORKERS) as executor:
        future_to_idx = {
            executor.submit(
                _generate_worker,
                prop["data"],
                target_visa,
                prop["style"],
                custom_instructions,
                model,
            ): i
            for i, prop in enumerate(properties)
        }
        done = 0
        total = len(future_to_idx)
        for future in as_completed(future_to_idx):
            i = future_to_idx[future]
            name = properties[i]["filename"]
            prop_num = properties[i].get("property_number", "")
            try:
                post = future.result()
                if prop_num:
                    post["html_content"] = _insert_property_number_to_table(
                        post.get("html_content", ""), prop_num
                    )
                blog_posts[i] = {
                    "filename": name,
                    "property_number": prop_num,
                    "post": post,
                }
            except Exception as e:
                errors.append(format_error_korean(e, name))
            done += 1
            progress.progress(done / total, text=f"[{done}/{total}] 생성 완료")

    progress.progress(1.0, text="✅ 생성 완료")
    blog_posts = [bp for bp in blog_posts if bp]

    if blog_posts:
        st.session_state["blog_posts"] = blog_posts
        # 디스크에 영구 이력 저장 (작업 이력 보관함에서 조회)
        new_history_items = []
        for bp in blog_posts:
            post = bp["post"]
            new_history_items.append({
                "filename": bp["filename"],
                "property_number": bp.get("property_number", ""),
                "title": post.get("title", ""),
                "summary_for_chat": post.get("summary_for_chat", ""),
                "html_content": post.get("html_content", ""),
                "hashtags": post.get("hashtags", []),
            })
        add_to_history(new_history_items, user_email=current_email)
        _persist_session()

        # 번역 DB 미등록 항목 수집 → 경고 표시
        all_untranslated = []
        for bp in blog_posts:
            for item in bp["post"].get("untranslated", []):
                all_untranslated.append(item)
        seen_keys = set()
        unique_untranslated = []
        for item in all_untranslated:
            k = (item.get("category"), item.get("original"))
            if k not in seen_keys:
                seen_keys.add(k)
                unique_untranslated.append(item)
        st.session_state["untranslated_alert"] = unique_untranslated

        st.success(
            f"✅ 블로그 {len(blog_posts)}개 생성 완료! "
            f"3번 탭에서 확인하시거나, **4️⃣ 작업 이력 보관함 탭**에서 "
            f"나중에라도 다시 조회할 수 있습니다."
        )
        st.balloons()
        # 2초 대기 후 새로고침 → 이력 보관함 자동 갱신
        time.sleep(2)
        st.rerun()
    if errors:
        st.error("⚠️ 일부 블로그 생성 실패")
        for err_msg in errors:
            st.markdown(err_msg)


# ─────────────────────────────────────────────────
# Drive 자동 처리 (Phase 3)
# ─────────────────────────────────────────────────

def _drive_configured() -> bool:
    """Drive 환경변수 두 개가 모두 설정돼 있는지 빠른 체크."""
    return bool(
        os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
        and os.getenv("DRIVE_ROOT_FOLDER_ID", "").strip()
    )


def _ensure_today_folder_cached():
    """오늘 폴더 ID/이름을 session_state에 10분 캐시. ⭐ 모든 시간은 도쿄 기준."""
    last_check = st.session_state.get("_today_folder_check_ts")
    cached = st.session_state.get("_today_folder")
    # 배포 직후 이전 세션의 naive datetime은 폐기 (tz-aware와 비교 시 TypeError 방지)
    if last_check and last_check.tzinfo is None:
        last_check = None
    if cached and last_check and (datetime.now(TZ_TOKYO) - last_check).total_seconds() < 600:
        return cached
    from src.drive_sync import get_or_create_today_folder
    folder_id, folder_name = get_or_create_today_folder()
    st.session_state["_today_folder"] = (folder_id, folder_name)
    st.session_state["_today_folder_check_ts"] = datetime.now(TZ_TOKYO)
    return (folder_id, folder_name)


def _refresh_drive_pending(folder_id, force=False):
    """미처리 파일 목록을 session_state에 10분 캐시. force=True면 즉시 갱신. ⭐ 모든 시간은 도쿄 기준."""
    last_check = st.session_state.get("_drive_files_check_ts")
    cached = st.session_state.get("_drive_pending_files")
    # 배포 직후 이전 세션의 naive datetime은 폐기
    if last_check and last_check.tzinfo is None:
        last_check = None
    if (
        not force
        and cached is not None
        and last_check
        and (datetime.now(TZ_TOKYO) - last_check).total_seconds() < 600
    ):
        return cached
    from src.drive_sync import list_pending_files
    pending = list_pending_files(folder_id)
    st.session_state["_drive_pending_files"] = pending
    st.session_state["_drive_files_check_ts"] = datetime.now(TZ_TOKYO)
    return pending


def _process_drive_files(file_list, current_email, target_visa_arg, model_arg, engine_arg):
    """Drive 자동 처리 파이프라인 — 다운로드 → 분석 → 생성 → 이력 저장 → 이동."""
    from src.drive_sync import (
        download_file_bytes,
        move_file_to_folder,
        get_or_create_processed_subfolder,
        append_auto_log,
    )

    # 1. 매물번호(7자리) 필터링 — 없는 파일은 자동 처리에서 스킵
    valid = []
    skipped = []
    for f in file_list:
        prop_num = _extract_property_number(f["name"])
        if not prop_num:
            skipped.append(f)
        else:
            valid.append({"file": f, "property_number": prop_num})

    if skipped:
        skipped_names = ", ".join(f["name"] for f in skipped)
        st.warning(
            f"⚠️ 매물번호(7자리) 없어서 자동 처리에서 제외 — **{len(skipped)}건**\n\n"
            f"파일명: {skipped_names}\n\n"
            f"→ Drive에서 파일명 앞에 7자리 매물번호 붙여주세요 "
            f"(예: `1234567_도면.pdf`)"
        )

    if not valid:
        st.info("ℹ️ 처리할 매물이 없습니다 (모든 파일이 매물번호 없음).")
        # 스킵만이라도 로그 기록
        if skipped:
            entries = [
                {
                    "timestamp": datetime.now(TZ_TOKYO).isoformat(),
                    "filename": f["name"],
                    "drive_file_id": f["id"],
                    "user_email": current_email,
                    "status": "skipped",
                    "error": "매물번호(7자리) 없음",
                }
                for f in skipped
            ]
            try:
                append_auto_log(entries)
            except Exception:
                pass
        return

    # 2. 처리완료 하위 폴더 확보
    try:
        folder_id, folder_name = st.session_state["_today_folder"]
        processed_folder_id = get_or_create_processed_subfolder(folder_id)
    except Exception as e:
        st.error(f"❌ '처리완료' 폴더 생성 실패: {e}")
        return

    # 3. 재처리 판정용 — 기존 이력의 파일명 set
    try:
        existing_history = load_history()
    except Exception:
        existing_history = []
    existing_filenames = {h.get("filename", "") for h in existing_history}

    st.markdown("---")
    st.markdown(f"### 🚀 Drive 자동 처리 — **{len(valid)}건**")

    # 4. 다운로드 + 분석 (병렬)
    dl_progress = st.progress(0.0, text="다운로드·분석 중…")

    def _dl_and_analyze(item):
        f = item["file"]
        bytes_data = download_file_bytes(f["id"])
        # ⭐ 파일명이 아닌 MIME 기반 확장자 사용 — Drive 사본 ("xxx.pdf의 사본") 대응
        # list_pending_files가 derived_ext 필드를 채워둠. 없으면 .pdf 기본값.
        suffix = f.get("derived_ext") or ".pdf"
        return _analyze_worker(bytes_data, suffix, engine_arg, model_arg)

    analyzed = []  # [{"file_info":..., "data":..., "error":...}]
    with ThreadPoolExecutor(max_workers=MAX_PARALLEL_WORKERS) as executor:
        future_to_idx = {
            executor.submit(_dl_and_analyze, item): i
            for i, item in enumerate(valid)
        }
        done = 0
        total = len(future_to_idx)
        for future in as_completed(future_to_idx):
            i = future_to_idx[future]
            item = valid[i]
            try:
                data = future.result()
                analyzed.append({"file_info": item, "data": data, "error": None})
            except Exception as e:
                analyzed.append({"file_info": item, "data": None, "error": format_error_korean(e, item["file"]["name"])})
            done += 1
            dl_progress.progress(done / total, text=f"[{done}/{total}] 다운로드·분석")

    # ⭐ 추출 결과를 物件DB(Postgres)에 저장 — 블로그 생성과 무관하게 먼저 저장.
    #    DB 미설정/실패해도 예외가 안 나므로 블로그 흐름엔 전혀 영향 없음.
    if property_db is not None:
        try:
            _db_items = [
                {
                    "property_number": a["file_info"]["property_number"],
                    "filename": a["file_info"]["file"]["name"],
                    "drive_file_id": a["file_info"]["file"]["id"],
                    "data": a["data"],
                    "source": "drive_auto",
                }
                for a in analyzed if a.get("data")
            ]
            if _db_items:
                property_db.save_extracted_batch(_db_items)
        except Exception:
            pass  # 物件DB 저장 실패는 블로그 흐름에 영향 주지 않음

    # 5. 분석 성공한 매물만 블로그 생성 (병렬)
    gen_progress = st.progress(0.0, text="블로그 생성 대기…")
    valid_for_gen = [a for a in analyzed if a.get("data")]

    successful = []  # 생성까지 성공한 매물 (이동 + 이력 저장 대상)
    failed = [a for a in analyzed if a.get("error")]  # 분석 실패한 것 먼저 추가

    if valid_for_gen:
        gen_progress.progress(0.0, text="블로그 생성 중…")
        with ThreadPoolExecutor(max_workers=MAX_PARALLEL_WORKERS) as executor:
            future_to_idx = {
                executor.submit(
                    _generate_worker,
                    a["data"],
                    target_visa_arg,
                    "친근형",  # 자동 모드 기본 스타일
                    "",        # 별도 지시 없음
                    model_arg,
                ): i
                for i, a in enumerate(valid_for_gen)
            }
            done = 0
            total = len(future_to_idx)
            for future in as_completed(future_to_idx):
                i = future_to_idx[future]
                a = valid_for_gen[i]
                try:
                    post = future.result()
                    a["post"] = post
                    successful.append(a)
                except Exception as e:
                    a["error"] = format_error_korean(e, a["file_info"]["file"]["name"])
                    failed.append(a)
                done += 1
                gen_progress.progress(done / total, text=f"[{done}/{total}] 블로그 생성")

    # 6. 성공한 매물: 이력 저장 + Drive 처리완료로 이동
    history_items_to_add = []
    log_entries = []
    moved_count = 0
    move_failed_names = []

    for s in successful:
        f = s["file_info"]["file"]
        prop_num = s["file_info"]["property_number"]
        post = s["post"]

        # 매물번호를 본문 표에 삽입 (수동 모드와 동일)
        if prop_num:
            post["html_content"] = _insert_property_number_to_table(
                post.get("html_content", ""), prop_num
            )

        # 재처리 판정 — 같은 파일명이 이미 이력에 있나?
        is_reprocess = f["name"] in existing_filenames

        history_items_to_add.append({
            "filename": f["name"],
            "property_number": prop_num,
            "title": post.get("title", ""),
            "summary_for_chat": post.get("summary_for_chat", ""),
            "html_content": post.get("html_content", ""),
            "hashtags": post.get("hashtags", []),
            "source": "drive_auto",     # 🤖 자동 배지 트리거
            "is_reprocess": is_reprocess,  # 🔄 재처리 배지 트리거
            "drive_file_id": f["id"],
        })

        # Drive 처리완료로 이동
        try:
            move_file_to_folder(f["id"], processed_folder_id)
            moved_count += 1
            log_entries.append({
                "timestamp": datetime.now(TZ_TOKYO).isoformat(),
                "filename": f["name"],
                "drive_file_id": f["id"],
                "user_email": current_email,
                "status": "success",
                "is_reprocess": is_reprocess,
            })
        except Exception as e:
            move_failed_names.append(f["name"])
            log_entries.append({
                "timestamp": datetime.now(TZ_TOKYO).isoformat(),
                "filename": f["name"],
                "drive_file_id": f["id"],
                "user_email": current_email,
                "status": "success_no_move",
                "error": str(e),
                "is_reprocess": is_reprocess,
            })

    # 이력 일괄 추가
    if history_items_to_add:
        try:
            add_to_history(history_items_to_add, user_email=current_email)
        except Exception as e:
            st.warning(f"⚠️ 이력 저장 중 오류: {e}")

    # 7. 실패·스킵 로그 기록
    for fl in failed:
        f = fl["file_info"]["file"]
        log_entries.append({
            "timestamp": datetime.now(TZ_TOKYO).isoformat(),
            "filename": f["name"],
            "drive_file_id": f["id"],
            "user_email": current_email,
            "status": "failed",
            "error": fl.get("error", "알 수 없는 오류"),
        })
    for sk in skipped:
        log_entries.append({
            "timestamp": datetime.now(TZ_TOKYO).isoformat(),
            "filename": sk["name"],
            "drive_file_id": sk["id"],
            "user_email": current_email,
            "status": "skipped",
            "error": "매물번호(7자리) 없음",
        })

    if log_entries:
        try:
            append_auto_log(log_entries)
        except Exception as e:
            st.warning(f"⚠️ 자동 처리 로그 저장 실패: {e}")

    # 8. 결과 요약
    gen_progress.progress(1.0, text="✅ 처리 완료")
    summary_lines = []
    if moved_count > 0:
        summary_lines.append(f"✅ **{moved_count}건 처리 완료** → '처리완료' 폴더로 이동")
    if move_failed_names:
        summary_lines.append(
            f"⚠️ {len(move_failed_names)}건 처리됐으나 Drive 이동 실패 (수동 이동 필요): "
            + ", ".join(move_failed_names)
        )
    if failed:
        summary_lines.append(f"❌ **{len(failed)}건 실패** (Drive 원위치 유지 → 재시도 가능)")

    if moved_count > 0:
        st.success("\n\n".join(summary_lines))
        st.balloons()
    elif summary_lines:
        st.warning("\n\n".join(summary_lines))

    if failed:
        with st.expander(f"❌ 실패 상세 ({len(failed)}건)", expanded=True):
            for fl in failed:
                f = fl["file_info"]["file"]
                st.markdown(f"- **{f['name']}**: {fl.get('error', '알 수 없음')}")

    # 9. 캐시 비우기 → 다음 진입 시 새로 폴링
    st.session_state.pop("_drive_pending_files", None)
    st.session_state.pop("_drive_files_check_ts", None)

    # 새로고침 (4번 탭 이력 자동 갱신)
    time.sleep(2)
    st.rerun()


def _render_drive_sync_area(current_email, target_visa_arg, model_arg, engine_arg):
    """1번 탭 상단의 Drive 자동 처리 영역 (사장님·사원 공통)."""
    if not _drive_configured():
        with st.expander("💡 Google Drive 자동 처리 (설정 필요)", expanded=False):
            st.caption(
                "관리자가 사이드바 '🔗 Drive 연결 테스트'에서 설정 후 사용 가능합니다."
            )
        return

    # 오늘 폴더 확보 (10분 캐시) — 사장님 요청 Q2-A
    try:
        folder_id, folder_name = _ensure_today_folder_cached()
    except Exception as e:
        st.error(
            f"❌ Drive '오늘 폴더({_today_folder_label()})' 확인 실패: {e}\n\n"
            "사이드바 '🔗 Drive 연결 테스트'로 권한을 다시 확인해주세요."
        )
        return

    # 새 파일 폴링 (10분 캐시)
    try:
        pending = _refresh_drive_pending(folder_id, force=False)
    except Exception as e:
        st.error(f"❌ Drive 파일 조회 실패: {e}")
        return

    last_check = st.session_state.get("_drive_files_check_ts")
    last_check_str = last_check.strftime("%H:%M") if last_check else "?"

    with st.container(border=True):
        # 헤더 + 다시 확인 버튼 (헤더 바로 옆에 가깝게 붙임)
        col1, col2, col_spacer = st.columns([3, 1.4, 4])
        with col1:
            st.markdown("### 🔄 Google Drive 자동 처리")
            st.caption(
                f"📁 오늘 폴더: **{folder_name}**  ·  "
                f"마지막 확인: {last_check_str}  ·  "
                f"10분마다 자동 갱신"
            )
        with col2:
            # 헤더에 시각적으로 더 붙도록 위쪽 여백 추가
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            if st.button(
                "🔄 지금 다시 확인",
                use_container_width=True,
                key="drive_refresh_btn",
            ):
                try:
                    _refresh_drive_pending(folder_id, force=True)
                except Exception as e:
                    st.error(f"❌ 갱신 실패: {e}")
                st.rerun()

        # 새 파일 표시
        if not pending:
            st.info("✨ 새 도면이 없습니다.")
        else:
            # 매물번호 유무 분리
            with_pn = [f for f in pending if _extract_property_number(f["name"])]
            without_pn = [f for f in pending if not _extract_property_number(f["name"])]

            if with_pn:
                st.success(f"🆕 **처리 대상 {len(with_pn)}건** 감지!")
            if without_pn:
                st.warning(
                    f"⚠️ 매물번호 없음 (스킵 대상) **{len(without_pn)}건**"
                )

            with st.expander(f"📋 파일 목록 ({len(pending)}건)", expanded=False):
                for f in pending:
                    name = f["name"]
                    size_mb = f.get("size", 0) / 1024 / 1024
                    prop_num = _extract_property_number(name)
                    if prop_num:
                        st.markdown(f"- ✅ `{name}` · {size_mb:.1f}MB · 매물번호 {prop_num}")
                    else:
                        st.markdown(f"- ⚠️ `{name}` · {size_mb:.1f}MB · 매물번호 없음 (스킵 예정)")

            # 처리 시작 버튼 (사장님·사원 동일)
            if with_pn:
                if st.button(
                    f"🚀 {len(with_pn)}건 자동 처리 시작 (분석 → 블로그 생성 → 이력 저장 → Drive 이동)",
                    type="primary",
                    use_container_width=True,
                    key="drive_process_btn",
                ):
                    _process_drive_files(
                        pending, current_email, target_visa_arg, model_arg, engine_arg
                    )

        # 자동 처리 로그 (최근 30일)
        with st.expander("📋 최근 자동 처리 로그 (30일 보관)", expanded=False):
            try:
                from src.drive_sync import load_auto_log
                log = load_auto_log()
            except ImportError:
                log = []
            if not log:
                st.caption("아직 자동 처리 기록이 없습니다.")
            else:
                # 최근 50건, 최신순
                recent = sorted(
                    log, key=lambda x: x.get("timestamp", ""), reverse=True
                )[:50]
                for entry in recent:
                    ts = entry.get("timestamp", "")
                    try:
                        ts_display = datetime.fromisoformat(ts).strftime("%m/%d %H:%M")
                    except (ValueError, TypeError):
                        ts_display = ts
                    status = entry.get("status", "")
                    fname = entry.get("filename", "?")
                    user = entry.get("user_email", "")
                    if status == "success":
                        icon = "✅ "
                    elif status == "success_no_move":
                        icon = "⚠️ "
                    elif status == "failed":
                        icon = "❌ "
                    elif status == "skipped":
                        icon = "⏭️ "
                    else:
                        icon = "• "
                    line = f"{icon} `{ts_display}` `{fname}` *({user})*"
                    if status in ("failed", "skipped", "success_no_move") and entry.get("error"):
                        err = entry["error"][:120]
                        line += f"  \n   ↳ {err}"
                    st.markdown(line)


def _today_folder_label() -> str:
    """오류 메시지용 도쿄 오늘 날짜 라벨."""
    try:
        from src.drive_sync import _today_folder_name
        return _today_folder_name()
    except Exception:
        return datetime.now(TZ_TOKYO).strftime("%Y%m%d")


def _render_naver_copy_button(html_content: str, unique_key: str) -> None:
    """
    네이버 발행용 서식 포함 복사 버튼 (Tab 3 미리보기·Tab 4 이력 보관함 공통).

    Args:
        html_content: 복사할 HTML (post.html_content)
        unique_key: 버튼·status DOM ID에 사용할 식별자 (Tab 3는 idx, Tab 4는 hid)
    """
    if not html_content:
        st.caption("⚠️ 본문 내용이 없어 복사 버튼을 표시할 수 없습니다.")
        return

    html_json = json.dumps(html_content)  # JS에 안전 전달

    st.components.v1.html(
        f"""
        <button id="naver-btn-{unique_key}"
            style="
                background:#03c75a;
                color:white;
                border:none;
                padding:14px 24px;
                font-size:15px;
                font-weight:600;
                border-radius:8px;
                cursor:pointer;
                width:100%;
                box-shadow:0 2px 4px rgba(0,0,0,0.1);
            ">
            📋 서식 포함 복사 (네이버 붙여넣기용)
        </button>
        <p id="status-{unique_key}" style="
            margin-top:8px;
            font-size:13px;
            color:#555;
            font-family:sans-serif;
            min-height:18px;
        "></p>
        <script>
            document.getElementById('naver-btn-{unique_key}').addEventListener('click', async function() {{
                const html = {html_json};
                const status = document.getElementById('status-{unique_key}');
                try {{
                    // ⭐ 서식 포함(rich) 복사 — 네이버 에디터에 바로 Ctrl+V 하면
                    //    표·체크리스트·이모지·인사말이 모양 그대로 들어감
                    const blobHtml = new Blob([html], {{ type: 'text/html' }});
                    const blobText = new Blob([html], {{ type: 'text/plain' }});
                    const item = new ClipboardItem({{
                        'text/html': blobHtml,
                        'text/plain': blobText
                    }});
                    await navigator.clipboard.write([item]);
                    status.innerHTML = '✅ 서식 포함 복사 완료! 네이버 글쓰기 화면에 바로 Ctrl+V';
                    status.style.color = '#03c75a';
                }} catch (err) {{
                    // 폴백: 일부 구형 브라우저는 ClipboardItem 미지원 → 텍스트로 복사
                    try {{
                        await navigator.clipboard.writeText(html);
                        status.innerHTML = '⚠️ 텍스트로만 복사됨 (이 브라우저는 서식 복사 미지원). Chrome 권장.';
                        status.style.color = '#e67e22';
                    }} catch (e2) {{
                        status.innerHTML = '❌ 복사 실패: ' + e2.message + ' (아래 HTML 다운로드 사용)';
                        status.style.color = '#d32f2f';
                    }}
                }}
            }});
        </script>
        """,
        height=110,
    )

    st.caption(
        "💡 **사용법**: 위 초록색 버튼 클릭(서식 포함 복사) → 네이버 블로그 글쓰기 "
        "화면에 **Ctrl+V** → 발행.  \n"
        "⚠️ 네이버 에디터는 붙여넣을 때 **정렬·글자 크기를 리셋**합니다 (네이버 정책). "
        "필요 시 **Ctrl+A 전체선택 → 가운데 정렬 → 글자 15** 한 번이면 글 전체에 적용됩니다."
    )


def _render_katok_copy_button(text_content: str, unique_key: str) -> None:
    """
    카카오톡 요약 텍스트 복사 버튼 — Tab 4 이력 보관함의 카톡 요약 모드용.
    원본 st.code() 영역의 자동 복사 아이콘이 우측 끝에 작아서 잘 안 보이는 문제 해결.

    Args:
        text_content: 복사할 텍스트 (summary_for_chat)
        unique_key: 버튼 DOM ID에 사용할 식별자
    """
    if not text_content:
        return

    text_json = json.dumps(text_content)  # JS에 안전 전달 (개행·특수문자 보존)

    st.components.v1.html(
        f"""
        <button id="katok-btn-{unique_key}"
            style="
                background:#FEE500;
                color:#3C1E1E;
                border:none;
                padding:12px 20px;
                font-size:15px;
                font-weight:600;
                border-radius:8px;
                cursor:pointer;
                width:100%;
                box-shadow:0 2px 4px rgba(0,0,0,0.08);
            ">
            💬 카톡 요약 복사 (붙여넣기용)
        </button>
        <p id="katok-status-{unique_key}" style="
            margin-top:6px;
            font-size:12px;
            color:#555;
            font-family:sans-serif;
            min-height:16px;
        "></p>
        <script>
            document.getElementById('katok-btn-{unique_key}').addEventListener('click', async function() {{
                const text = {text_json};
                const status = document.getElementById('katok-status-{unique_key}');
                try {{
                    await navigator.clipboard.writeText(text);
                    status.innerHTML = '✅ 복사 완료! 카카오톡에 Ctrl+V';
                    status.style.color = '#03c75a';
                }} catch (err) {{
                    status.innerHTML = '❌ 복사 실패: ' + err.message;
                    status.style.color = '#d32f2f';
                }}
            }});
        </script>
        """,
        height=85,
    )


# ─────────────────────────────────────────────────
# 4단계 탭
# ─────────────────────────────────────────────────

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(
    [
        "1️⃣ 도면 업로드",
        "2️⃣ 추출 결과·스타일 선택",
        "3️⃣ 블로그 미리보기",
        "4️⃣ 📚 작업 이력 보관함",
        "5️⃣ 📊 통계",
        "6️⃣ 🏠 物件DB·제안",
    ]
)

# ───── Tab 1: 업로드 (병렬 분석) ─────
with tab1:
    # ──────────────────────────────────────────────
    # 🔄 Google Drive 자동 처리 영역 (Phase 3)
    # 사장님·사원 누구나 접근 가능. 새 파일 폴링 + [🚀 처리 시작]
    # ──────────────────────────────────────────────
    _render_drive_sync_area(current_email, target_visa, model, engine)

    st.divider()

    # 💡 Drive 안 거치고 직접 업로드 (예외 상황용) — 평소엔 접혀 있음
    with st.expander("💡 Drive 안 거치고 직접 업로드 (예외 상황용)", expanded=False):
        st.subheader(f"마이소크 (物件図面) 직접 업로드 — 한 번에 최대 {MAX_UPLOADS}개")
        st.caption("지원 형식: JPG · PNG · WEBP · GIF · PDF  ·  Drive 안 거치고 즉시 처리할 때 사용")
        st.info(
            "📌 **매물번호 안내**: 파일명 맨 앞에 **7자리 숫자**를 붙여주세요.\n\n"
            "예: `1234567_매물도면.jpg` → 매물번호 **1234567**\n\n"
            "매물번호는 블로그 제목·이력 보관함·검색에 사용됩니다."
        )

        uploaded_files = st.file_uploader(
            f"도면 파일을 최대 {MAX_UPLOADS}개까지 선택하세요",
            type=["jpg", "jpeg", "png", "webp", "gif", "pdf"],
            accept_multiple_files=True,
        )

        if uploaded_files:
            if len(uploaded_files) > MAX_UPLOADS:
                st.warning(f"⚠️ 최대 {MAX_UPLOADS}개까지만 처리됩니다.")
                uploaded_files = uploaded_files[:MAX_UPLOADS]

            # 매물번호 미인식 파일 경고
            no_number_files = [
                uf.name for uf in uploaded_files
                if not _extract_property_number(uf.name)
            ]
            if no_number_files:
                st.warning(
                    "⚠️ **다음 파일은 7자리 매물번호가 인식되지 않았습니다:**\n\n"
                    + "\n".join(f"- {n}" for n in no_number_files)
                    + "\n\n매물번호 없이도 진행되지만, 파일명 앞에 7자리 숫자를 "
                    "붙이는 것을 권장합니다. (예: `1234567_원래파일명.jpg`)"
                )

            st.write(f"**업로드된 파일: {len(uploaded_files)}개**")
            cols = st.columns(min(len(uploaded_files), MAX_UPLOADS))
            for i, uf in enumerate(uploaded_files):
                with cols[i]:
                    if uf.name.lower().endswith(".pdf"):
                        st.info(f"📄 {uf.name}\n(PDF)")
                    else:
                        st.image(uf, caption=uf.name, use_container_width=True)

            # ⭐ 자동 모드: 분석 완료 후 블로그 생성을 자동으로 이어서 진행
            # 기본값을 OFF로 변경 (켜져 있으면 4번 탭이 35-70초간 비활성)
            # 새 탭에서 이력 보관함을 따로 보고 싶으면 옆 버튼 사용
            _col_auto, _col_newtab = st.columns([3, 2])
            with _col_auto:
                auto_generate_enabled = st.checkbox(
                    "🚀 자동 모드 — 분석 후 블로그까지 한 번에 생성",
                    value=False,
                    key="auto_generate_enabled",
                    help=(
                        "켜면 분석 후 블로그 생성까지 35~70초 끊김 없이 진행. "
                        "그동안 4번 탭(이력 보관함)이 비활성화됩니다. "
                        "처리 중에도 이력을 보고 싶으면 오른쪽 [새 탭에서 열기] 버튼 사용."
                    ),
                )
            with _col_newtab:
                _app_url = os.getenv("APP_BASE_URL", "https://jreblog.onrender.com")
                st.link_button(
                    "🪟 새 탭에서 이력 보관함 열기",
                    _app_url,
                    help="처리 중에도 다른 탭에서 작업 이력 자유롭게 사용 가능",
                    use_container_width=True,
                )

            if st.button("🔍 전체 도면 병렬 분석 시작", type="primary"):
                # ⭐ 이전 분석 결과 완전 초기화 (다른 도면인데 같은 결과 나오는 버그 방지)
                st.session_state.pop("properties", None)
                st.session_state.pop("blog_posts", None)
                st.session_state.pop("untranslated_alert", None)

                properties = []
                errors = []
                progress = st.progress(0.0, text="병렬 분석 시작…")

                # 파일 미리 읽기 (Streamlit UploadedFile은 thread-safe 안 함)
                # ⭐ 파일명 중복 방지: 같은 이름이면 인덱스 부여
                file_jobs = []
                seen_names = {}
                for uf in uploaded_files:
                    base_name = uf.name
                    if base_name in seen_names:
                        seen_names[base_name] += 1
                        # 같은 파일명 구분 (확장자 앞에 _2, _3)
                        stem = Path(base_name).stem
                        suf = Path(base_name).suffix
                        unique_name = f"{stem}_{seen_names[base_name]}{suf}"
                    else:
                        seen_names[base_name] = 1
                        unique_name = base_name
                    file_jobs.append((unique_name, uf.getvalue(), Path(uf.name).suffix))

                with ThreadPoolExecutor(max_workers=MAX_PARALLEL_WORKERS) as executor:
                    # ⭐ 인덱스 기반 매핑 (파일명 중복돼도 결과 안 섞임)
                    future_to_idx = {
                        executor.submit(_analyze_worker, file_bytes, suffix, engine, model): idx
                        for idx, (name, file_bytes, suffix) in enumerate(file_jobs)
                    }

                    results_by_idx = {}
                    done = 0
                    total = len(future_to_idx)
                    for future in as_completed(future_to_idx):
                        idx = future_to_idx[future]
                        name = file_jobs[idx][0]
                        try:
                            data = future.result()
                            results_by_idx[idx] = {
                                "filename": name,
                                "property_number": _extract_property_number(name),
                                "data": data,
                                "style": default_style,  # 기본 스타일
                            }
                        except Exception as e:
                            errors.append(format_error_korean(e, name))
                        done += 1
                        progress.progress(
                            done / total,
                            text=f"[{done}/{total}] 분석 완료",
                        )

                progress.progress(1.0, text="✅ 분석 완료")
                # 업로드 순서대로 정렬 (인덱스 순)
                properties = [results_by_idx[i] for i in sorted(results_by_idx.keys())]

                # ⭐ 추출 결과를 物件DB(Postgres)에 저장 (직접 업로드는 drive_file_id 없음)
                #    DB 미설정/실패해도 예외가 안 나므로 흐름엔 영향 없음.
                if property_db is not None and properties:
                    try:
                        property_db.save_extracted_batch([
                            {
                                "property_number": p.get("property_number", ""),
                                "filename": p.get("filename", ""),
                                "drive_file_id": None,
                                "data": p.get("data") or {},
                                "source": "upload",
                            }
                            for p in properties
                        ])
                    except Exception:
                        pass

                if properties:
                    st.session_state["properties"] = properties
                    st.session_state.pop("blog_posts", None)
                    # 이 세션에서 분석을 했음을 표시 (복원이 덮어쓰지 못하게)
                    st.session_state["_analysis_done_this_session"] = True
                    # 디스크에 자동 저장 (새 결과로 덮어쓰기)
                    _persist_session(overwrite=True)

                    # ⭐ 자동 모드: 분석 직후 블로그 생성을 같은 클릭에서 이어서 진행
                    if auto_generate_enabled:
                        st.success(
                            f"✅ {len(properties)}개 도면 분석 완료. 블로그 자동 생성을 시작합니다…"
                        )
                        st.markdown("---")
                        st.markdown("### ✍️ 블로그 자동 생성 중")
                        # 자동 모드는 custom_instructions 없이 진행 (필요 시 자동 모드 끄고 수동)
                        _run_blog_generation(
                            properties=properties,
                            custom_instructions="",
                            target_visa=target_visa,
                            model=model,
                            current_email=current_email,
                        )
                    else:
                        st.success(
                            f"✅ {len(properties)}개 도면 분석 완료! 2번 탭에서 확인하세요."
                        )
                if errors:
                    st.error("⚠️ 일부 파일 분석 실패")
                    for err_msg in errors:
                        st.markdown(err_msg)

# ───── Tab 2: 추출 결과 + 파일별 스타일 선택 ─────
with tab2:
    properties = st.session_state.get("properties")
    if not properties:
        st.info("👈 먼저 1번 탭에서 도면을 업로드하고 분석을 실행하세요.")
    else:
        st.subheader(f"📋 추출 결과 — 총 {len(properties)}개")
        st.caption(
            "AI 추출 정보를 확인하고, 매물별로 글 스타일을 선택하세요. "
            "틀린 정보는 직접 수정 가능합니다."
        )

        for idx, prop in enumerate(properties):
            data = prop["data"]
            station = data.get("nearest_station") or {}
            # 제목: 파일명(일본어 PDF명) 대신 핵심 정보로 표시 + 관리비 포함
            _rent = data.get("rent_yen", 0) or 0
            _mgmt = data.get("management_fee_yen", 0) or 0
            _layout = data.get("layout", "?")
            if _mgmt > 0:
                _price_label = f"월세 ¥{_rent:,} + 관리비 ¥{_mgmt:,}"
            else:
                _price_label = f"월세 ¥{_rent:,}"
            _prop_num = prop.get("property_number", "")
            _num_label = f"[{_prop_num}] " if _prop_num else ""
            with st.expander(
                f"📄 {idx+1}. {_num_label}{_layout} / {_price_label}",
                expanded=(idx == 0),
            ):
                st.caption(f"📁 원본 파일: {prop['filename']}")
                col1, col2, col3 = st.columns([2, 2, 1])
                with col1:
                    st.markdown(
                        f"**가장 가까운 역**: {station.get('line', '?')} "
                        f"{station.get('station', '?')}역 "
                        f"도보 {station.get('walk_minutes', '?')}분"
                    )
                    st.markdown(
                        f"**방구조 / 전용면적**: {data.get('layout', '?')} "
                        f"/ {data.get('area_sqm', '?')}㎡"
                    )
                with col2:
                    mgmt = data.get("management_fee_yen") or 0
                    st.metric(
                        "월세",
                        f"¥{data.get('rent_yen', 0):,}",
                        f"관리비 ¥{mgmt:,}",
                        delta_color="off",
                    )
                with col3:
                    conf = data.get("extraction_confidence", "?")
                    if conf == "low":
                        st.error(f"⚠️ 자신도: {conf}")
                    elif conf == "medium":
                        st.warning(f"자신도: {conf}")
                    else:
                        st.success(f"자신도: {conf}")
                    engine_used = data.get("_engine_used", "?")
                    if "gemini" in engine_used and "claude" not in engine_used:
                        st.caption(f"🆓 {engine_used}")
                    elif "폴백" in engine_used:
                        st.caption(f"🔀 {engine_used}")
                    else:
                        st.caption(f"💎 {engine_used}")

                # 이 매물의 스타일 선택 (파일별 다른 스타일 가능)
                prop["style"] = st.selectbox(
                    f"이 매물의 글 스타일",
                    options=available_styles,
                    index=available_styles.index(prop.get("style", default_style))
                    if prop.get("style", default_style) in available_styles
                    else 0,
                    key=f"style_select_{idx}",
                    help="매물마다 다른 스타일을 선택할 수 있습니다.",
                )

                edited = st.text_area(
                    "추출 데이터 (필요시 직접 수정)",
                    value=json.dumps(data, ensure_ascii=False, indent=2),
                    height=240,
                    key=f"json_edit_{idx}",
                )
                try:
                    prop["data"] = json.loads(edited)
                except json.JSONDecodeError as e:
                    st.warning(f"⚠️ JSON 형식 오류: {e}")

        st.divider()
        st.markdown("### ✍️ 블로그 글 일괄 생성 (병렬 처리)")

        custom_instructions = st.text_area(
            "전체 글 공통 특별 지시 (선택)",
            placeholder=(
                "전체 글에 공통 적용할 지시. 예:\n"
                "• 여성 손님 대상, 안전성 강조\n"
                "• 한인 마트·한국 음식점 정보 비중 늘리기"
            ),
            height=80,
        )

        # 스타일 요약 표시
        style_summary = ", ".join(
            f"{i+1}번: {p['style']}" for i, p in enumerate(properties)
        )
        st.caption(f"📝 선택된 스타일: {style_summary}")

        if st.button(f"✍️ 블로그 {len(properties)}개 일괄 생성", type="primary"):
            # ⭐ 공통 헬퍼 호출 (Tab 1 자동 모드와 동일 로직)
            _run_blog_generation(
                properties=properties,
                custom_instructions=custom_instructions,
                target_visa=target_visa,
                model=model,
                current_email=current_email,
            )

# ───── Tab 3: 블로그 미리보기 ─────
with tab3:
    blog_posts = st.session_state.get("blog_posts")

    if not blog_posts:
        st.info(
            "👈 2번 탭에서 블로그 글을 생성하세요.\n\n"
            "💡 과거에 생성한 블로그 글은 **4️⃣ 작업 이력 보관함 탭**에서 다시 조회·다운로드 가능합니다."
        )
    else:
        st.subheader(f"📝 생성된 블로그 — 총 {len(blog_posts)}개")

        # ⭐ DB(路線 시트) 미등록 항목 경고 — 회사 DB 추가 등록 안내
        _untrans = st.session_state.get("untranslated_alert", [])
        if _untrans:
            warn_lines = []
            for item in _untrans:
                warn_lines.append(
                    f"- **{item.get('category')}**: `{item.get('original')}` "
                    f"({item.get('note', '')})"
                )
            st.warning(
                "⚠️ **회사 번역 DB에 등록되지 않은 항목이 있습니다.**\n\n"
                + "\n".join(warn_lines)
                + "\n\n위 항목은 일본어가 한국어로 번역되지 않았을 수 있습니다. "
                "**路線 시트(번역 DB)에 추가 등록**하면 다음부터 자동 번역됩니다.\n\n"
                "👉 DB: https://docs.google.com/spreadsheets/d/1D6u75qwjPodXS82SWaZhJ0MzNIn3Hf_GkPMiYutZthA"
            )

        # 전체 ZIP 다운로드 (G드라이브 저장용)
        target_folder = OUTPUT_FOLDER_PATH
        timestamp_label = datetime.now().strftime("%Y년%m월%d일 %H시%M분")
        timestamp_file = datetime.now().strftime("%Y%m%d_%H%M")

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for idx, bp in enumerate(blog_posts):
                post = bp["post"]
                stem = Path(bp["filename"]).stem
                html = build_naver_smarteditor_html(post)
                zf.writestr(f"{timestamp_file}_{idx+1:02d}_{stem}.html", html)
                zf.writestr(
                    f"{timestamp_file}_{idx+1:02d}_{stem}.json",
                    json.dumps(post, ensure_ascii=False, indent=2),
                )
        zip_buf.seek(0)

        # 큰 안내 박스
        st.markdown(
            f"""
            <div style="background:#e3f2fd;border-left:5px solid #1976d2;
                        padding:14px 18px;border-radius:6px;margin:12px 0">
                <div style="font-size:15px;font-weight:600;color:#0d47a1;
                            margin-bottom:6px">
                    💾 작업 결과 저장 안내 ({timestamp_label} 작업분 — {len(blog_posts)}개)
                </div>
                <div style="font-size:13px;color:#1a1a2e;line-height:1.7">
                    아래 <b>ZIP 다운로드 버튼</b>을 누르면 압축 파일이 생성됩니다.<br>
                    다운로드 창에서 <b>저장 위치를 다음 폴더로 지정</b>하세요:
                </div>
                <div style="background:#fff;border:1px solid #bbdefb;
                            padding:8px 12px;border-radius:4px;margin-top:8px;
                            font-family:'Courier New',monospace;font-size:12px;
                            color:#0d47a1;word-break:break-all">
                    {target_folder}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        col_a, col_b = st.columns([1, 1])
        with col_a:
            # 경로 한 줄 → 우측 📋 아이콘으로 클립보드 복사
            st.markdown("**📋 저장 경로 (우측 아이콘 클릭하면 복사됨)**")
            st.code(target_folder, language=None)
        with col_b:
            st.markdown("**📦 ZIP 다운로드**")
            st.download_button(
                f"📦 ZIP 다운로드 ({len(blog_posts)}개 블로그)",
                zip_buf.getvalue(),
                file_name=f"블로그_{timestamp_file}.zip",
                mime="application/zip",
                type="primary",
                use_container_width=True,
            )

        with st.expander("💡 매번 자동으로 G드라이브 폴더에 저장되게 하는 방법"):
            st.markdown(
                f"""
                **Chrome 다운로드 위치를 G드라이브 폴더로 설정하면 매번 자동 저장됩니다.**

                1. Chrome 우측 위 **⋮ → 설정**
                2. 왼쪽 메뉴 **다운로드**
                3. **"위치"** 옆 **"변경"** 클릭
                4. 다음 폴더 선택:
                   ```
                   {target_folder}
                   ```
                5. **"다운로드 전에 각 파일의 저장 위치 확인"** → 꺼두기

                > ⚠️ 이 설정 후에는 다른 곳에서 받는 파일도 이 폴더로 갑니다.
                > 매번 위치를 묻게 하려면 5번을 켜두세요. (그래도 시작 위치가
                > G드라이브 폴더라 매번 클릭 두세 번이면 끝납니다.)
                """
            )

        st.divider()

        for idx, bp in enumerate(blog_posts):
            post = bp["post"]
            with st.expander(
                f"📝 {idx+1}. {post.get('title', bp['filename'])}",
                expanded=(idx == 0),
            ):
                st.text_input("제목", value=post.get("title", ""), key=f"title_{idx}")

                # ⭐ 카카오톡용 요약 — 큰 노란색 복사 버튼 (눈에 띄게)
                st.markdown("**📱 카카오톡용 요약**")
                _render_katok_copy_button(post.get("summary_for_chat", ""), f"tab3_{idx}")
                with st.expander("👁 요약 내용 미리보기", expanded=False):
                    st.code(post.get("summary_for_chat", ""), language=None, wrap_lines=True)

                st.markdown("**미리보기**")
                st.caption("🚨 빨간색 배경의 '⚠️ 현지 확인 필요' 부분은 직접 채워 넣으세요")
                st.markdown(post.get("html_content", ""), unsafe_allow_html=True)

                st.markdown("**해시태그**")
                st.code(" ".join(post.get("hashtags", [])), language=None)

                # ⭐ 네이버 발행 보조 — 본문 HTML 클립보드 복사 (공통 헬퍼 사용)
                st.markdown("---")
                st.markdown("**✍️ 네이버 블로그 발행 보조**")
                _render_naver_copy_button(post.get("html_content", ""), str(idx))

                c1, c2 = st.columns(2)
                with c1:
                    st.download_button(
                        "💾 HTML 다운로드 (백업)",
                        build_naver_smarteditor_html(post),
                        file_name=f"blog_{idx+1}_{Path(bp['filename']).stem}.html",
                        mime="text/html",
                        key=f"dl_html_{idx}",
                    )
                with c2:
                    st.download_button(
                        "📋 JSON 다운로드",
                        json.dumps(post, ensure_ascii=False, indent=2),
                        file_name=f"blog_{idx+1}.json",
                        mime="application/json",
                        key=f"dl_json_{idx}",
                    )

# ───── Tab 4: 작업 이력 보관함 (구 카페발행 탭 위치) ─────
with tab4:
    try:
        _hist_preview = load_history()
        # 작업 이력은 모든 사용자가 공유 (관리자/사용자 구분 없음)
        _hist_count = len(_hist_preview)
    except Exception:
        _hist_preview = []
        _hist_count = 0

    # 파일 크기 계산 (모든 사용자에게 표시)
    try:
        from src.persistence import HISTORY_FILE
        if HISTORY_FILE.exists():
            _hist_bytes = HISTORY_FILE.stat().st_size
            if _hist_bytes < 1024:
                _hist_size = f"{_hist_bytes}B"
            elif _hist_bytes < 1024 * 1024:
                _hist_size = f"{_hist_bytes/1024:.1f}KB"
            else:
                _hist_size = f"{_hist_bytes/(1024*1024):.2f}MB"
        else:
            _hist_size = "0B"
    except Exception:
        _hist_size = ""

    # 최근 작업 / 가장 오래된 계산
    _hist_latest = ""
    _hist_oldest = ""
    try:
        if _hist_preview:
            _latest_ts = datetime.fromisoformat(_hist_preview[0]["timestamp"])
            _hist_latest = _latest_ts.strftime("%m/%d %H:%M")
            _oldest_ts = datetime.fromisoformat(_hist_preview[-1]["timestamp"])
            _days_old = (datetime.now() - _oldest_ts).days
            _hist_oldest = f"{_days_old}일전" if _days_old > 0 else "오늘"
    except Exception:
        pass

    # 한 줄 통계+안내 (상단 공간 최소화)
    if _hist_count == 0:
        st.caption("🗂️ 비어 있음 · 1·2번 탭에서 블로그 생성 시 자동 저장됨")
    else:
        _parts = [f"{_hist_count}건"]
        if _hist_size:
            _parts.append(_hist_size)
        if _hist_latest:
            _parts.append(f"🆕{_hist_latest}")
        if _hist_oldest:
            _parts.append(f"🗓{_hist_oldest}")
        st.caption(
            f"🗂️ {' · '.join(_parts)}  ·  영구 보존, 회사 전체 공유"
        )

    # 이력 로드 (실패해도 빈 리스트 반환) — 모든 사용자가 전체 이력 조회
    try:
        history_all = load_history()
        # 기존 이력 (user_email 없음) → admin@win-bro.com 작성으로 처리
        for h in history_all:
            if not h.get("user_email"):
                h["user_email"] = ADMIN_EMAIL
        # 작업 이력은 회사 공유: 모든 사용자가 전체 이력 조회
        history = history_all
    except Exception as e:
        st.error(f"⚠️ 이력 로드 중 에러: {e}")
        history = []
        history_all = []

    if not history:
        st.info(
            "아직 저장된 이력이 없습니다.\n\n"
            "1·2번 탭에서 블로그를 생성하면 자동으로 여기에 저장됩니다."
        )
    else:
        # 검색 + 즐겨찾기 필터 — 검색창 짧게, 즐겨찾기 왼쪽으로
        _col_search, _col_fav, _col_blank = st.columns([2, 1, 3])
        with _col_search:
            search_q = st.text_input(
                "🔍 검색 (매물번호·제목·파일명)",
                key="hist_search_v2",
                placeholder="예: 1234567, 신주쿠, 1K",
                label_visibility="collapsed",
            )
        with _col_fav:
            only_favorites = st.checkbox(
                "⭐ 즐겨찾기만",
                key="hist_only_favorites_v2",
                value=False,
            )
        # _col_blank는 의도적으로 빈 공간

        # 필터링 (매물번호·제목·파일명)
        filtered = history
        if search_q:
            q = search_q.lower()
            filtered = [
                h for h in filtered
                if q in h.get("title", "").lower()
                or q in h.get("filename", "").lower()
                or q in h.get("property_number", "").lower()
            ]
        if only_favorites:
            filtered = [h for h in filtered if h.get("favorite")]

        if not filtered:
            st.warning(f"'{search_q}' 검색 결과 없음")
        else:
            # 일괄 작업 영역 — 전체/해제/선택삭제를 왼쪽편(블로그 제목 위)에 배치
            # 카운트는 오른쪽으로
            # selected_ids는 아래 루프 후에 확정되지만, 이전 rerun의 session_state로 미리 계산
            _pre_selected = [
                h["id"] for h in filtered
                if st.session_state.get(f"hist_sel_{h['id']}")
            ]
            _col_sel, _col_unsel, _col_del, _col_spacer, _col_cnt = st.columns(
                [1.2, 1.2, 1.8, 2, 2.5]
            )
            with _col_sel:
                if st.button("☑️ 전체", key="hist_select_all", use_container_width=True):
                    for h in filtered:
                        st.session_state[f"hist_sel_{h['id']}"] = True
                    st.rerun()
            with _col_unsel:
                if st.button("⬜ 해제", key="hist_unselect_all", use_container_width=True):
                    for h in filtered:
                        st.session_state[f"hist_sel_{h['id']}"] = False
                    st.rerun()
            with _col_del:
                _del_label = (
                    f"🗑️ 삭제 ({len(_pre_selected)})"
                    if _pre_selected
                    else "🗑️ 삭제"
                )
                if st.button(
                    _del_label,
                    key="hist_delete_top",
                    type="primary" if _pre_selected else "secondary",
                    disabled=(len(_pre_selected) == 0),
                    use_container_width=True,
                    help="체크한 항목 삭제",
                ):
                    delete_from_history(_pre_selected)
                    for sid in _pre_selected:
                        st.session_state.pop(f"hist_sel_{sid}", None)
                    st.success(f"✅ {len(_pre_selected)}개 항목 삭제 완료")
                    st.rerun()
            # _col_spacer는 빈 공간
            with _col_cnt:
                st.caption(f"검색 결과 **{len(filtered)}건**")

            # 이력 목록 표시
            selected_ids = []
            for idx, h in enumerate(filtered):
                hid = h.get("id", "")
                title = h.get("title", "(제목 없음)")
                filename = h.get("filename", "")
                prop_num = h.get("property_number", "")
                timestamp = h.get("timestamp", "")
                is_fav = h.get("favorite", False)

                # 제목에 이미 [매물번호]가 포함됐는지 확인 (중복 방지)
                title_has_num = prop_num and title.startswith(f"[{prop_num}]")

                # 표시용 시간 포맷
                try:
                    ts_obj = datetime.fromisoformat(timestamp)
                    ts_display = ts_obj.strftime("%Y-%m-%d %H:%M")
                    # ⭐ 영구 보존이라 자동 삭제 경고 없음
                    retention_warn = ""
                except Exception:
                    ts_display = timestamp
                    retention_warn = ""

                # 체크박스 + 즐겨찾기 + 제목
                col_chk, col_fav, col_info = st.columns([0.5, 0.5, 9])
                with col_chk:
                    checked = st.checkbox(
                        " ",
                        key=f"hist_sel_{hid}",
                        label_visibility="collapsed",
                    )
                    if checked:
                        selected_ids.append(hid)

                with col_fav:
                    fav_icon = "⭐" if is_fav else "☆"
                    if st.button(
                        fav_icon,
                        key=f"hist_fav_{hid}",
                        help="즐겨찾기 토글",
                        use_container_width=True,
                    ):
                        toggle_favorite(hid)
                        st.rerun()

                with col_info:
                    fav_badge = " ⭐" if is_fav else ""
                    # 🤖 자동 처리 / 🔄 재처리 배지 (Phase 3)
                    source = h.get("source", "manual")
                    is_reprocess = h.get("is_reprocess", False)
                    auto_badges = []
                    if source == "drive_auto":
                        auto_badges.append("🤖")
                    if is_reprocess:
                        auto_badges.append("🔄")
                    badge_prefix = ("".join(auto_badges) + " ") if auto_badges else ""
                    # 순번 다음에 매물번호 표시 (제목에 이미 있으면 제목만)
                    if prop_num and not title_has_num:
                        display_title = f"**{idx+1}. {badge_prefix}[{prop_num}] {title}**{fav_badge}"
                    else:
                        display_title = f"**{idx+1}. {badge_prefix}{title}**{fav_badge}"
                    st.markdown(
                        f"{display_title}  \n"
                        f"<span style='color:#888;font-size:13px'>"
                        f"📁 {filename} · 🕒 {ts_display}{retention_warn}</span>",
                        unsafe_allow_html=True,
                    )

                # 📂 상세 보기 — 카톡 요약/본문 HTML/다운로드를 한 expander 안에 라디오로 통합
                summary = h.get("summary_for_chat", "")
                html = h.get("html_content", "")
                tags = h.get("hashtags", [])

                _mode_options = []
                if summary:
                    _mode_options.append("📱 카톡 요약")
                if html:
                    _mode_options.append("📄 본문 HTML")
                _mode_options.append("💾 다운로드")

                with st.expander("📂 상세 보기", expanded=False):
                    _mode = st.radio(
                        "보기 모드",
                        _mode_options,
                        horizontal=True,
                        key=f"hist_mode_{hid}",
                        label_visibility="collapsed",
                    )

                    if _mode == "📱 카톡 요약":
                        # ⭐ 큰 노란색 복사 버튼 (눈에 띄게)
                        _render_katok_copy_button(summary, f"hist_{hid}")
                        with st.expander("👁 요약 내용 미리보기", expanded=False):
                            st.code(summary, language=None, wrap_lines=True)
                    elif _mode == "📄 본문 HTML":
                        if html:
                            # ⭐ 옵션 A: 네이버 발행용 서식 포함 복사 버튼
                            # (Tab 3 안 거치고 Tab 4에서 바로 발행 가능 — Drive 자동 처리 후 일상 흐름)
                            st.markdown("**✍️ 네이버 블로그 발행 보조**")
                            _render_naver_copy_button(html, f"hist_{hid}")
                            st.markdown("**미리보기**")
                            st.markdown(html, unsafe_allow_html=True)
                        if tags:
                            st.markdown("**해시태그**")
                            st.code(" ".join(tags), language=None)
                    elif _mode == "💾 다운로드":
                        dl_col1, dl_col2 = st.columns(2)
                        with dl_col1:
                            st.download_button(
                                "💾 HTML 다운로드",
                                build_naver_smarteditor_html({
                                    "title": title,
                                    "html_content": html,
                                    "hashtags": tags,
                                }),
                                file_name=f"history_{hid}_{Path(filename).stem}.html",
                                mime="text/html",
                                key=f"hist_dl_html_{hid}",
                                use_container_width=True,
                            )
                        with dl_col2:
                            st.download_button(
                                "📋 JSON 다운로드",
                                json.dumps(h, ensure_ascii=False, indent=2),
                                file_name=f"history_{hid}.json",
                                mime="application/json",
                                key=f"hist_dl_json_{hid}",
                                use_container_width=True,
                            )

                # 얇은 회색선 (st.divider 대신 — 위아래 큰 패딩 제거)
                st.markdown(
                    '<hr style="border:none;border-top:1px solid #eee;margin:6px 0">',
                    unsafe_allow_html=True,
                )

            # 하단: 전체 삭제 + 백업 다운로드 (선택 삭제는 상단으로 이동됨)
            st.markdown("---")
            col_d2, col_d3 = st.columns([1, 2])
            with col_d2:
                if st.button(
                    "🗑️ 전체 삭제",
                    type="secondary",
                    use_container_width=True,
                ):
                    if st.session_state.get("_confirm_clear_all"):
                        clear_history()
                        st.session_state.pop("_confirm_clear_all", None)
                        st.success("✅ 모든 이력 삭제 완료")
                        st.rerun()
                    else:
                        st.session_state["_confirm_clear_all"] = True
                        st.warning("⚠️ 한 번 더 클릭하면 모든 이력이 삭제됩니다.")

            with col_d3:
                # 전체 백업 다운로드
                if filtered:
                    st.download_button(
                        f"📥 전체 백업 다운로드 ({len(history)}건 JSON)",
                        json.dumps(history, ensure_ascii=False, indent=2),
                        file_name=f"jre_history_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                        mime="application/json",
                        use_container_width=True,
                    )




# ───── Tab 5: 📊 통계 (회사 전체 + 직원별, sub-tabs) ─────
with tab5:
    st.markdown("### 📊 작업 통계")
    st.caption("회사 전체 통계와 직원별 작업 현황을 한눈에 확인할 수 있습니다.")
    st.divider()

    _stats_tab_a, _stats_tab_b = st.tabs(
        ["📊 회사 전체 통계", "📈 직원별 작업 통계"]
    )
    with _stats_tab_a:
        _render_company_stats()
    with _stats_tab_b:
        _render_staff_stats()


# ───── Tab 6: 🏠 物件DB · 제안 리스트 (전체 칼럼·직접편집 표) ─────
with tab6:
    st.markdown("### 🏠 物件DB · 제안 리스트")
    st.caption("전체 칼럼을 한 표에서 보고, 사진링크·시키킹·레이킹은 표에서 직접 편집할 수 있습니다. (좌우로 스크롤)")

    if property_db is None or not property_db.is_configured():
        st.warning(
            "物件DB(데이터베이스)가 연결되지 않았습니다. "
            "환경변수 DATABASE_URL을 확인해 주세요."
        )
    else:
        import pandas as pd

        # 상단: 새로고침(왼쪽) + 검색(오른쪽)
        _top_l, _top_r = st.columns([1, 3])
        with _top_l:
            if st.button("🔄 새로고침", key="propdb_refresh", use_container_width=True):
                st.rerun()
        with _top_r:
            _q = st.text_input(
                "검색", key="propdb_search",
                placeholder="🔍 건물명·주소·매물번호·역·지역 검색",
                label_visibility="collapsed",
            )

        # 관리 줄: 모집종료 포함 보기 + 2주 지난 매물 정리
        _mng_l, _mng_r = st.columns([1, 1])
        with _mng_l:
            _show_closed = st.checkbox("모집종료 포함 보기", key="propdb_show_closed")
        with _mng_r:
            if st.button("🧹 2주 지난 매물 정리", key="propdb_cleanup", use_container_width=True):
                if st.session_state.get("_propdb_confirm_cleanup"):
                    try:
                        _n = property_db.delete_old(14)
                        st.session_state.pop("_propdb_confirm_cleanup", None)
                        st.success(f"2주 지난 매물 {_n}건을 정리했습니다.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"정리 실패: {e}")
                else:
                    st.session_state["_propdb_confirm_cleanup"] = True
                    st.warning("⚠️ 한 번 더 누르면 2주 지난 매물이 영구 삭제됩니다.")

        _props = []
        try:
            _props = property_db.list_properties(
                include_closed=bool(st.session_state.get("propdb_show_closed")),
                limit=500,
            )
        except Exception as e:
            st.error(f"매물 목록을 불러오지 못했습니다: {e}")

        # 복합 검색 필터 (건물명·주소·매물번호·역·지역, 공백으로 여러 단어 AND)
        if _q and _q.strip():
            _terms = _q.strip().lower().split()

            def _match(p):
                d = p.get("data") or {}
                stn = (d.get("nearest_station") or {}).get("station") or ""
                hay = " ".join([
                    str(p.get("property_number") or ""),
                    str(d.get("property_name") or ""),
                    str(d.get("address") or ""),
                    str(stn),
                    _ward_ko(d.get("address") or ""),
                ]).lower()
                return all(t in hay for t in _terms)

            _props = [p for p in _props if _match(p)]

        st.caption(
            f"총 매물 {len(_props)}건"
            + (f" · 검색: '{_q.strip()}'" if (_q and _q.strip()) else " · 매물번호 최신순")
        )

        if not _props:
            st.info("표시할 매물이 없습니다. (검색어를 지우거나 도면을 추출해 보세요)")
        else:
            # 행 구성 + 매물 id 매핑(저장용) + 원본 행 전체 보관(변경 감지용)
            _rows, _id_list, _orig_rows = [], [], []
            for _p in _props:
                _row = _property_to_full_row(_p)
                _rows.append(_row)
                _id_list.append(_p.get("id"))
                _orig_rows.append(dict(_row))

            _df = pd.DataFrame(_rows, columns=_FULL_COLS)
            _df.insert(0, "선택", False)  # 맨 앞 체크박스 칼럼

            _colcfg = {
                "선택": st.column_config.CheckboxColumn("선택", help="블로그·제안·모집종료에 쓸 매물 체크"),
                "맵": st.column_config.LinkColumn("맵", display_text="집위치보기"),
                "매물검색": st.column_config.LinkColumn("매물검색", display_text="검색결과"),
                "사진링크": st.column_config.TextColumn("사진링크", help="중개사이트 매물 링크 붙여넣기"),
                "시키킹": st.column_config.NumberColumn("시키킹(개월분)", min_value=0.0, step=0.5, help="개월수 입력 (예: 1.0). 0이면 없음"),
                "레이킹": st.column_config.NumberColumn("레이킹(개월분)", min_value=0.0, step=0.5, help="개월수 입력 (예: 1.0). 0이면 없음"),
            }

            st.caption("모든 칸을 직접 편집할 수 있습니다. 빈칸도 입력 가능. 칼럼 머리글 클릭=정렬. 수정 후 아래 저장 버튼을 누르세요.")
            _edited = st.data_editor(
                _df,
                key="propdb_editor",
                use_container_width=True,
                hide_index=True,
                column_config=_colcfg,
                height=420,
            )

            if st.button("💾 변경사항 저장", key="propdb_save_all", type="primary"):
                _saved, _failed = 0, 0

                def _mk_dep(months):
                    try:
                        v = float(months)
                    except (ValueError, TypeError):
                        return None
                    return {"value": v, "unit": "months"} if v > 0 else None

                for _i in range(len(_id_list)):
                    _rid = _id_list[_i]
                    _orig = _orig_rows[_i]

                    # 변경된 칼럼 수집
                    _diff = {}
                    for _col in _FULL_COLS:
                        _a = _edited.iloc[_i][_col]
                        _b = _orig.get(_col)
                        if _col in _DEPOSIT_COLS:
                            try:
                                if abs(float(_a or 0) - float(_b or 0)) > 1e-9:
                                    _diff[_col] = _a
                            except (ValueError, TypeError):
                                if _a != _b:
                                    _diff[_col] = _a
                        else:
                            if str(_a if _a is not None else "").strip() != str(_b if _b is not None else "").strip():
                                _diff[_col] = _a
                    if not _diff:
                        continue

                    try:
                        _cur = property_db.get_property(_rid) or {}
                        _manual = dict(_cur.get("manual_fields") or {})
                        _ov = dict(_manual.get("overrides") or {})

                        for _col, _val in _diff.items():
                            if _col == "사진링크":
                                _s = str(_val or "").strip()
                                if _s:
                                    _manual["photo_link"] = _s
                                else:
                                    _manual.pop("photo_link", None)
                            elif _col == "시키킹":
                                _d = _mk_dep(_val)
                                _manual["shikikin"] = _d if _d else None
                                if _d is None:
                                    _manual.pop("shikikin", None)
                            elif _col == "레이킹":
                                _d = _mk_dep(_val)
                                _manual["reikin"] = _d if _d else None
                                if _d is None:
                                    _manual.pop("reikin", None)
                            else:
                                _s = str(_val if _val is not None else "").strip()
                                if _s:
                                    _ov[_col] = _s
                                else:
                                    _ov.pop(_col, None)

                        if _ov:
                            _manual["overrides"] = _ov
                        else:
                            _manual.pop("overrides", None)

                        property_db.update_manual_fields(_rid, _manual)
                        _saved += 1
                    except Exception:
                        _failed += 1

                if _saved:
                    st.success(f"{_saved}건 저장되었습니다." + (f" ({_failed}건 실패)" if _failed else ""))
                    st.rerun()
                elif _failed:
                    st.error(f"{_failed}건 저장 실패")
                else:
                    st.info("변경된 내용이 없습니다.")

            # ── 선택 매물 액션 (4-3a) ──
            st.divider()
            _sel_idx = [i for i in range(len(_id_list)) if bool(_edited.iloc[i]["선택"])]
            _sel_ids = [_id_list[i] for i in _sel_idx]
            st.markdown(f"#### ✅ 선택한 매물: {len(_sel_idx)}건")

            _a1, _a2, _a3 = st.columns(3)
            with _a1:
                if st.button("📋 선택 매물 모아보기", key="propdb_collect", use_container_width=True,
                             disabled=(len(_sel_idx) == 0)):
                    st.session_state["_propdb_show_selected"] = True
            with _a2:
                if st.button("🚫 모집종료 처리", key="propdb_close", use_container_width=True,
                             disabled=(len(_sel_idx) == 0)):
                    _c, _e = 0, 0
                    for _rid in _sel_ids:
                        try:
                            property_db.mark_closed(_rid, True)
                            _c += 1
                        except Exception:
                            _e += 1
                    if _c:
                        st.success(f"{_c}건 모집종료 처리했습니다." + (f" ({_e}건 실패)" if _e else ""))
                        st.rerun()
                    elif _e:
                        st.error(f"{_e}건 처리 실패")
            with _a3:
                st.button("📝 블로그/제안 만들기 (다음 단계)", key="propdb_make",
                          use_container_width=True, disabled=True,
                          help="블로그 생성·손님용 제안 링크는 다음 단계에서 연결됩니다.")

            # 선택 매물 모아보기 (제안 리스트 기초 — 핵심 칼럼만)
            if st.session_state.get("_propdb_show_selected") and _sel_idx:
                st.markdown("##### 선택한 매물 모아보기")
                _brief_cols = [
                    "매물번호", "건물명", "지역", "월세", "관리비", "시키킹", "레이킹",
                    "노선1", "가까운역1", "도보1", "신주쿠까지", "간취", "면적", "입주일", "비자",
                ]
                _brief = _edited.iloc[_sel_idx][[c for c in _brief_cols if c in _edited.columns]]
                st.dataframe(_brief, use_container_width=True, hide_index=True)
                st.caption("※ 손님용 제안 카드·링크는 5단계에서 만듭니다. 지금은 선택 내용 확인용입니다.")
